import os
import sys
# 添加项目根目录到Python路径，解决模块导入问题
from src.utils.project_path import get_project_root
project_root = get_project_root()
sys.path.insert(0, project_root)


try:
    from .feature_extractor.feature_extractor import FeatureExtractor
except ImportError:
    from src.evaluator.feature_extractor.feature_extractor import FeatureExtractor

# 导入CBG链接生成器
try:
    from ..utils.cbg_link_generator import CBGLinkGenerator
    from ..utils import load_jsonc_from_config_dir
except ImportError:
    try:
        from utils.cbg_link_generator import CBGLinkGenerator
        from utils import load_jsonc_from_config_dir
    except ImportError:
        import sys
        import os
        sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'utils'))
        from cbg_link_generator import CBGLinkGenerator
        from jsonc_loader import load_jsonc_from_config_dir


import sqlite3
import json
import logging
import numpy as np
from typing import Dict,  Optional, Union
import glob
import pandas as pd



class RuleEvaluator:
    def __init__(self):
    
        self.logger = logging.getLogger(__name__)
        self.feature_extractor = FeatureExtractor()
        self.df = None

    def _extract_features(self, character_data, conn=None):
        """提取角色特征

        使用 FeatureExtractor 提取特征
        """
        try:
            # 使用 FeatureExtractor 提取特征
            features = self.feature_extractor.extract_features(character_data)
            return features

        except Exception as e:
            self.logger.error(f"特征提取失败: {e}")
            raise

    def calc_appearance_value(self, features):
        """计算外观系统价值"""
        value = 0

        # 限量锦衣价值
        if 'limited_skin_value' in features:
            # 指数衰减模型：前几件价值高，后面边际递减
            skin_value = features['limited_skin_value']
            value += 1000 * (1 - np.exp(-skin_value / 500))

        return value

    def evaluate(self, character_data):
        """评估角色价值
        
        Args:
            character_data: 角色数据
        """
        try:
            # 1.提取特征
            features = self._extract_features(character_data)
            
            # 使用改进的算法
            improved_result = self.calc_base_attributes_value_improved(features)
            
            evaluation_result = {
                'features': features,
                'base_value': improved_result['final_value'],
                'value_breakdown': improved_result['value_breakdown'],
                'total_raw_value': improved_result['total_raw_value'],
                'algorithm_version': 'improved'
            }

            return evaluation_result

        except Exception as e:
            self.logger.error(f"角色评估失败: {e}")
            raise

    def batch_evaluate(self, db_path):
        """加载数据"""
        try:
            print(f"正在连接数据库: {db_path}")
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()

            # 获取所有角色数据
            print("正在查询数据库...")
            cursor.execute("""
                SELECT 
                    c.id, c.equip_id, c.server_name, c.seller_nickname, c.level, c.price,
                    c.price_desc, c.school AS school_desc, c.area_name, c.icon_index,
                    c.kindid, c.game_ordersn, c.pass_fair_show, c.fair_show_end_time,
                    c.accept_bargain, c.status_desc, c.onsale_expire_time_desc, c.expire_time,
                    c.race, c.fly_status, c.collect_num, c.life_skills, c.school_skills,
                    c.ju_qing_skills,c.yushoushu_skill, c.all_pets_json, c.all_equip_json AS all_equip_json_desc,
                    c.all_shenqi_json, c.all_rider_json AS all_rider_json_desc, c.all_fabao_json,
                    c.ex_avt_json AS ex_avt_json_desc, c.create_time, c.update_time,
                    l.*
                FROM characters c
                LEFT JOIN large_equip_desc_data l ON c.equip_id = l.equip_id
                WHERE c.price > 0 AND l.all_equip_json =='{}' AND l.all_summon_json == '[]'
            """)

            # 获取列名
            columns = [description[0] for description in cursor.description]
            print(f"查询完成，开始获取数据...")
            rows = cursor.fetchall()
            print(f"获取到 {len(rows)} 条数据")
            # 准备训练数据
            print("开始提取特征...")
            data = []
            for i, row in enumerate(rows):
                character_data = dict(zip(columns, row))
                features = self._extract_features(character_data, conn)
                features['price'] = character_data.get('price', 0)
                features['equip_id'] = character_data.get('equip_id', '')
                data.append(features)
                if (i + 1) % 10 == 0:
                    print(f"已处理 {i + 1}/{len(rows)} 条数据")

            print("特征提取完成，转换为DataFrame...")
            # 转换为DataFrame
            self.df = pd.DataFrame(data)
            print(f"DataFrame 形状: {self.df.shape}")
            self.df.set_index('equip_id', inplace=True)

            print("数据加载和预处理完成")

        except Exception as e:
            print(f"数据加载失败: {str(e)}")
            import traceback
            print(traceback.format_exc())
            raise
        finally:
            if 'conn' in locals():
                conn.close()

    def generate_cbg_link(self, eid: str) -> Optional[str]:

        return CBGLinkGenerator.generate_cbg_link(eid)

    def calc_base_attributes_value_improved(self, features: Dict[str, Union[int, float, str]]) -> Dict[str, float]:
        """
        简化的改进基础属性价值计算，只引入市场折价系数
        
        Args:
            features: 角色特征字典
            
        Returns:
            Dict[str, float]: 包含各项价值详情的字典
        """
        try:
            # 加载规则配置
            RULE = load_jsonc_from_config_dir('rule_setting.jsonc', __file__)
            
            # 加载折价率配置
            DISCOUNT_RATES = load_jsonc_from_config_dir('rate.jsonc', __file__)
            
            value_breakdown = {}
            total_value = 0
            
            # 按照原始算法计算各项价值，然后应用折价系数
            # 1. 历史门派数加成
            school_history_count = features.get('school_history_count', 0)
            raw_value = school_history_count * RULE.get('ChangeSchoolCostPerTimes', 0)
            discount_value = raw_value * DISCOUNT_RATES.get('school_history', 1.0)
            if raw_value > 0:
                value_breakdown['school_history'] = discount_value
                total_value += discount_value
                print(f"历史门派数: 原始{raw_value:.0f}万 MHB -> 市场{discount_value/RULE.get('RMB2MHB',15)*RULE.get('MARKET_FACTOR'):.0f}元 (系数{DISCOUNT_RATES.get('school_history', 1.0)})")
            
            # 2. 乾元丹加成
            all_new_point = features.get('all_new_point', 0)
            qianyuandan_dict = RULE.get('QianyuandanCount2Value', {})
            if str(all_new_point) in qianyuandan_dict:
                raw_value = all_new_point * qianyuandan_dict[str(all_new_point)]
                discount_value = raw_value * DISCOUNT_RATES.get('qianyuandan', 1.0)
                value_breakdown['qianyuandan'] = discount_value
                total_value += discount_value
                print(f"乾元丹: 原始{raw_value:.0f}万 MHB -> 市场{discount_value/RULE.get('RMB2MHB',15)*RULE.get('MARKET_FACTOR'):.0f}元 (系数{DISCOUNT_RATES.get('qianyuandan', 1.0)})")
            
            # 3. 乾元丹突破加成
            if features.get('qianyuandan_breakthrough', False):
                raw_value = RULE.get('qianyuandanBreakthroughValue', 0)
                discount_value = raw_value * DISCOUNT_RATES.get('qianyuandan', 1.0)  # 使用乾元丹的系数
                value_breakdown['qianyuandan_breakthrough'] = discount_value
                total_value += discount_value
                print(f"乾元丹突破: 原始{raw_value:.0f}万 MHB -> 市场{discount_value/RULE.get('RMB2MHB',15)*RULE.get('MARKET_FACTOR'):.0f}元 (系数{DISCOUNT_RATES.get('qianyuandan', 1.0)})")
            
            # 4. 月饼粽子机缘
            jiyuan_amount = features.get('jiyuan_amount', 0)
            raw_value = jiyuan_amount * RULE.get('JiyuanValuePerCount', 0)
            discount_value = raw_value * DISCOUNT_RATES.get('jiyuan', 1.0)
            if raw_value > 0:
                value_breakdown['jiyuan'] = discount_value
                total_value += discount_value
                print(f"月饼粽子机缘: 原始{raw_value:.0f}万 MHB -> 市场{discount_value/RULE.get('RMB2MHB',15)*RULE.get('MARKET_FACTOR'):.0f}元 (系数{DISCOUNT_RATES.get('jiyuan', 1.0)})")
            
            # 5. 总经验
            sum_exp = features.get('sum_exp', 0)
            raw_value = sum_exp * RULE.get('ExpValuePerHundredMillion', 0)
            discount_value = raw_value * DISCOUNT_RATES.get('exp', 1.0)
            if raw_value > 0:
                value_breakdown['exp'] = discount_value
                total_value += discount_value
                print(f"总经验: 原始{raw_value:.0f}万 MHB -> 市场{discount_value/RULE.get('RMB2MHB',15)*RULE.get('MARKET_FACTOR'):.0f}元 (系数{DISCOUNT_RATES.get('exp', 1.0)})")
            
            # 6. 师门技能
            school_skills = features.get('school_skills', [])
            school_skill_dict = RULE.get('SchoolSkillGrade2Value', {})
            raw_value = 0
            for skill_grade in school_skills:
                if str(skill_grade) in school_skill_dict:
                    raw_value += school_skill_dict[str(skill_grade)]
            discount_value = raw_value * DISCOUNT_RATES.get('school_skills', 1.0)
            if raw_value > 0:
                value_breakdown['school_skills'] = discount_value
                total_value += discount_value
                print(f"师门技能: 原始{raw_value:.0f}万 MHB -> 市场{discount_value/RULE.get('RMB2MHB',15)*RULE.get('MARKET_FACTOR'):.0f}元 (系数{DISCOUNT_RATES.get('school_skills', 1.0)})")
            
            # 7. 行囊
            packet_page = features.get('packet_page', 0)
            packet_dict = RULE.get('PackagePage2Value', {})
            if str(packet_page) in packet_dict:
                raw_value = packet_dict[str(packet_page)]
                discount_value = raw_value * DISCOUNT_RATES.get('packet', 1.0)
                value_breakdown['packet'] = discount_value
                total_value += discount_value
                print(f"行囊: 原始{raw_value:.0f}万 MHB -> 市场{discount_value/RULE.get('RMB2MHB',15)*RULE.get('MARKET_FACTOR'):.0f}元 (系数{DISCOUNT_RATES.get('packet', 1.0)})")
            
            # 8. 储备金
            learn_cash = features.get('learn_cash', 0)
            raw_value = learn_cash/10000 * RULE.get('LearnCash2CashRate', 0)
            discount_value = raw_value * DISCOUNT_RATES.get('learn_cash', 1.0)
            if raw_value > 0:
                value_breakdown['learn_cash'] = discount_value
                total_value += discount_value
                print(f"储备金: 原始{raw_value:.0f}万 MHB -> 市场{discount_value/RULE.get('RMB2MHB',15)*RULE.get('MARKET_FACTOR'):.0f}元 (系数{DISCOUNT_RATES.get('learn_cash', 1.0)})")
            
            # 9. 仙玉
            xianyu_amount = features.get('xianyu_amount', 0)
            raw_value = xianyu_amount * RULE.get('RMB2MHB', 0)
            discount_value = raw_value * DISCOUNT_RATES.get('xianyu', 1.0)
            if raw_value > 0:
                value_breakdown['xianyu'] = discount_value
                total_value += discount_value
                print(f"仙玉: 原始{raw_value:.0f}万 MHB -> 市场{discount_value/RULE.get('RMB2MHB',15)*RULE.get('MARKET_FACTOR'):.0f}元 (系数{DISCOUNT_RATES.get('xianyu', 1.0)})")
            
            # 10. 修炼等级上限
            max_cultivate_dict = RULE.get('MaxCultivateGrade2Value', {})
            cultivate_cost_dict = RULE.get('CultivateCostPerCount', {})
            raw_value = 0
            for i in range(1, 5):
                max_expt_key = f'max_expt{i}'
                max_expt_value = features.get(max_expt_key, 0)
                cost_multiplier = cultivate_cost_dict.get(str(i), 0)
                if max_expt_value > 0 and str(max_expt_value) in max_cultivate_dict:
                    raw_value += max_cultivate_dict[str(max_expt_value)] * cost_multiplier
            discount_value = raw_value * DISCOUNT_RATES.get('cultivation', 1.0)
            if raw_value > 0:
                value_breakdown['cultivation_limit'] = discount_value
                total_value += discount_value
                print(f"修炼等级上限: 原始{raw_value:.0f}万 MHB -> 市场{discount_value/RULE.get('RMB2MHB',15)*RULE.get('MARKET_FACTOR'):.0f}元 (系数{DISCOUNT_RATES.get('cultivation', 1.0)})")
            
            # 11. 修炼等级
            cultivate_learn_count_dict = RULE['CultivateGrade2LearnCount']
            raw_value = 0
            for i in range(1, 5):
                expt_key = f'expt_ski{i}'
                expt_value = features.get(expt_key, 0)
                cost_multiplier = cultivate_cost_dict.get(str(i), 0)
                raw_value += cultivate_learn_count_dict[str(expt_value)] * cost_multiplier
            discount_value = raw_value * DISCOUNT_RATES.get('cultivation', 1.0)
            if raw_value > 0:
                value_breakdown['cultivation_level'] = discount_value
                total_value += discount_value
                print(f"修炼等级: 原始{raw_value:.0f}万 MHB -> 市场{discount_value/RULE.get('RMB2MHB',15)*RULE.get('MARKET_FACTOR'):.0f}元 (系数{DISCOUNT_RATES.get('cultivation', 1.0)})")
            
            # 12. 控制力
            control_dict = RULE.get('ControlGrade2XiulianguoCount', {})
            xiulianguo_value = RULE.get('XiulianguoValuePerCount', 0)
            raw_value = 0
            for i in range(1, 5):
                beast_ski_key = f'beast_ski{i}'
                beast_ski_value = features.get(beast_ski_key, 0)
                if str(beast_ski_value) in control_dict:
                    raw_value += xiulianguo_value * control_dict[str(beast_ski_value)]
            discount_value = raw_value * DISCOUNT_RATES.get('beast_cultivation', 1.0)
            if raw_value > 0:
                value_breakdown['beast_cultivation'] = discount_value
                total_value += discount_value
                print(f"控制力: 原始{raw_value:.0f}万 MHB -> 市场{discount_value/RULE.get('RMB2MHB',15)*RULE.get('MARKET_FACTOR'):.0f}元 (系数{DISCOUNT_RATES.get('beast_cultivation', 1.0)})")
            
            # 13. 生活技能
            life_skills = features.get('life_skills', [])
            life_skill_dict = RULE.get('LifeSkillGrade2Value', {})
            raw_value = 0
            for life_skill_grade in life_skills:
                if life_skill_grade >= 80 and str(life_skill_grade) in life_skill_dict:
                    raw_value += life_skill_dict[str(life_skill_grade)]
            discount_value = raw_value * DISCOUNT_RATES.get('life_skills', 1.0)
            if raw_value > 0:
                value_breakdown['life_skills'] = discount_value
                total_value += discount_value
                print(f"生活技能: 原始{raw_value:.0f}万 MHB -> 市场{discount_value/RULE.get('RMB2MHB',15)*RULE.get('MARKET_FACTOR'):.0f}元 (系数{DISCOUNT_RATES.get('life_skills', 1.0)})")
            
            # 14. 强壮、神速
            qiangzhuang_shensu = features.get('qiangzhuang&shensu', [])
            qiangzhuang_dict = RULE.get('QiangzhuangAndShensuGrade2Value', {})
            raw_value = 0
            for qiangzhuang_and_shensu_grade in qiangzhuang_shensu:
                if str(qiangzhuang_and_shensu_grade) in qiangzhuang_dict:
                    raw_value += qiangzhuang_dict[str(qiangzhuang_and_shensu_grade)]
            discount_value = raw_value * DISCOUNT_RATES.get('qiangzhuang_shensu', 1.0)
            if raw_value > 0:
                value_breakdown['qiangzhuang_shensu'] = discount_value
                total_value += discount_value
                print(f"强壮、神速: 原始{raw_value:.0f}万 MHB -> 市场{discount_value/RULE.get('RMB2MHB',15)*RULE.get('MARKET_FACTOR'):.0f}元 (系数{DISCOUNT_RATES.get('qiangzhuang_shensu', 1.0)})")
            
            # 15. 灵佑次数
            lingyou_count = features.get('lingyou_count', 0)
            raw_value = lingyou_count * RULE.get('LingyouValuePerCount', 0)
            discount_value = raw_value * DISCOUNT_RATES.get('lingyou', 1.0)
            if raw_value > 0:
                value_breakdown['lingyou'] = discount_value
                total_value += discount_value
                print(f"灵佑次数: 原始{raw_value:.0f}万 MHB -> 市场{discount_value/RULE.get('RMB2MHB',15)*RULE.get('MARKET_FACTOR'):.0f}元 (系数{DISCOUNT_RATES.get('lingyou', 1.0)})")
            
            # 16. 高成长坐骑数量
            hight_grow_rider_count = features.get('hight_grow_rider_count', 0)
            raw_value = hight_grow_rider_count * RULE.get('HightGrowRidePerValue', 0)
            discount_value = raw_value * DISCOUNT_RATES.get('rider', 1.0)
            if raw_value > 0:
                value_breakdown['rider'] = discount_value
                total_value += discount_value
                print(f"高成长坐骑: 原始{raw_value:.0f}万 MHB -> 市场{discount_value/RULE.get('RMB2MHB',15)*RULE.get('MARKET_FACTOR'):.0f}元 (系数{DISCOUNT_RATES.get('rider', 1.0)})")
            
            # 17. 召唤兽最大携带量
            allow_pet_count = features.get('allow_pet_count', 0)
            raw_value = RULE.get('AllowPetCount2Value', {}).get(str(allow_pet_count), 0)
            discount_value = raw_value * DISCOUNT_RATES.get('pet_count', 1.0)
            if raw_value > 0:
                value_breakdown['pet_count'] = discount_value
                total_value += discount_value
                print(f"召唤兽携带量: 原始{raw_value:.0f}万 MHB -> 市场{discount_value/RULE.get('RMB2MHB',15)*RULE.get('MARKET_FACTOR'):.0f}元 (系数{DISCOUNT_RATES.get('pet_count', 1.0)})")
            
            # 18. 有价值法宝数量
            premium_fabao_count = features.get('premium_fabao_count', 0)
            raw_value = premium_fabao_count * RULE.get('PremiumFabaoValuePerCount', 0)
            discount_value = raw_value * DISCOUNT_RATES.get('fabao', 1.0)
            if raw_value > 0:
                value_breakdown['fabao'] = discount_value
                total_value += discount_value
                print(f"有价值法宝: 原始{raw_value:.0f}万 MHB -> 市场{discount_value/RULE.get('RMB2MHB',15)*RULE.get('MARKET_FACTOR'):.0f}元 (系数{DISCOUNT_RATES.get('fabao', 1.0)})")
            
            # 19. 神器
            shenqi = features.get('shenqi', [0,0,0,0,0,0,0,0])
            ShenqiAttr2Value = RULE.get('ShenqiAttr2Value', {})
            raw_value = 0
            # shenqi数组索引: [allTheSameCount, same9Count, 3attrCount, 3Count, 2attrCount, 2Count, 1attrCount, 1Count]
            shenqi_mapping = [
                (0, 'allTheSameCount'),  # allTheSameCount
                (1, 'same9Count'),       # same9Count
                (2, '3attr'),           # 3attrCount
                (3, '3'),               # 3Count
                (4, '2attr'),           # 2attrCount
                (5, '2'),               # 2Count
                (6, '1attr'),           # 1attrCount
                (7, '1')                # 1Count
            ]
            
            for index, multiplier_key in shenqi_mapping:
                if index < len(shenqi) and shenqi[index] > 0:
                    raw_value += shenqi[index] * ShenqiAttr2Value.get(multiplier_key, 0)
            discount_value = raw_value * DISCOUNT_RATES.get('shenqi', 1.0)
            if raw_value > 0:
                value_breakdown['shenqi'] = discount_value
                total_value += discount_value
                print(f"神器: 原始{raw_value:.0f}万 MHB -> 市场{discount_value/RULE.get('RMB2MHB',15)*RULE.get('MARKET_FACTOR'):.0f}元 (系数{DISCOUNT_RATES.get('shenqi', 1.0)})")
            
            # 20. 限量锦衣价值
            limited_skin_value = features.get('limited_skin_value', 0)
            if limited_skin_value > 0:
                # 直接以元为单位计入
                value_breakdown['limited_skin'] = limited_skin_value
                total_value += 0  # 不再计入MHB
                print(f"限量锦衣: 直接计入 {limited_skin_value} 元")
            
            # 21. 限量祥瑞价值
            limited_huge_horse_value = features.get('limited_huge_horse_value', 0)
            if limited_huge_horse_value > 0:
                # 直接以元为单位计入
                value_breakdown['limited_huge_horse'] = limited_huge_horse_value
                total_value += 0  # 不再计入MHB
                print(f"限量祥瑞: 直接计入 {limited_huge_horse_value} 元")
            
            # 最终价值转换
            final_value = total_value / RULE['RMB2MHB'] * RULE['MARKET_FACTOR']
            # 直接加上限量锦衣和祥瑞的元价值
            final_value += limited_skin_value + limited_huge_horse_value
            
            print(f"\n=== 简化版价值分解 ===")
            for attr, value in value_breakdown.items():
                print(f"{attr}: {value:.0f}")
            print(f"总计: {total_value:.0f} -> 最终: {final_value:.0f}")
            
            return {
                'value_breakdown': value_breakdown,
                'total_raw_value': total_value,
                'final_value': final_value
            }
            
        except Exception as e:
            self.logger.error(f"简化版改进基础属性价值计算失败: {e}")
            import traceback
            self.logger.error(f"详细错误: {traceback.format_exc()}")
            return {
                'value_breakdown': {},
                'total_raw_value': 0.0,
                'final_value': 0.0
            }

    def export_comparison_to_excel(self, output_path: str = 'results_comparison.xlsx') -> Optional[str]:
        """
        导出改进算法结果到Excel文件

        Args:
            output_path (str): 输出文件路径，默认为 'results_comparison.xlsx'

        Returns:
            Optional[str]: 成功时返回输出文件路径，失败时返回None
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows

            if self.df is None or self.df.empty:
                print("没有数据可导出")
                return None

            print(f"开始导出 {len(self.df)} 条数据到Excel（改进算法结果）...")

            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "改进算法结果"

            # 准备数据
            export_data = []
            for equip_id, row in self.df.iterrows():
                # 获取角色基本信息
                price = row.get('price', 0)

                # 计算改进算法价值
                try:
                    improved_result = self.calc_base_attributes_value_improved(row.to_dict())
                    improved_value = improved_result['final_value']
                    value_breakdown = improved_result['value_breakdown']
                except:
                    improved_value = 0
                    value_breakdown = {}

                # 生成CBG链接
                cbg_link = self.generate_cbg_link(equip_id)

                # 计算差异
                improved_diff_ratio = (improved_value - price) / price * 100 if price > 0 else 0

                # 准备一行数据
                row_data = {
                    'equip_id': equip_id,
                    'cbg_link': cbg_link,
                    'price': price,
                    'improved_value': improved_value,
                    'improved_diff_ratio': improved_diff_ratio,
                    **value_breakdown,  # 展开价值分解
                    **row.to_dict()  # 包含所有features
                }
                export_data.append(row_data)

            if not export_data:
                print("没有有效数据可导出")
                return None

            # 定义列顺序：优先列 + 价值分解列 + features列
            priority_columns = ['equip_id', 'price', 'improved_value', 'improved_diff_ratio']
            
            # 获取所有价值分解列
            value_breakdown_columns = []
            for data in export_data:
                for key in data.keys():
                    # 检查是否是价值分解列
                    if (key in ['school_history', 'jiyuan', 'exp', 'learn_cash', 'xianyu', 
                               'lingyou', 'rider', 'fabao', 'qianyuandan', 'packet', 'pet_count',
                               'qianyuandan_breakthrough', 'school_skills', 'cultivation_limit',
                               'cultivation_level', 'beast_cultivation', 'life_skills', 'qiangzhuang_shensu', 
                               'shenqi', 'limited_skin', 'limited_huge_horse']):
                        if key not in value_breakdown_columns:
                            value_breakdown_columns.append(key)

            # 获取所有features列（排除优先列和价值分解列）
            all_columns = list(export_data[0].keys())
            feature_columns = [
                col for col in all_columns 
                if col not in priority_columns and col != 'cbg_link' and col not in value_breakdown_columns
            ]

            # 最终列顺序
            final_columns = priority_columns + value_breakdown_columns + feature_columns

            # 设置表头
            headers = []
            for col in final_columns:
                if col == 'equip_id':
                    headers.append('装备ID')
                elif col == 'price':
                    headers.append('实际价格')
                elif col == 'improved_value':
                    headers.append('改进估值')
                elif col == 'improved_diff_ratio':
                    headers.append('差价比率(%)')
                elif col in value_breakdown_columns:
                    headers.append(f'{col}_价值')
                else:
                    headers.append(col)

            # 写入表头
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(
                    horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

            # 写入数据
            for row_idx, data in enumerate(export_data, 2):
                for col_idx, col_name in enumerate(final_columns, 1):
                    value = data.get(col_name, '')

                    # 特殊处理角色名列（添加超链接）
                    if col_name == 'equip_id':
                        cbg_link = data.get('cbg_link')
                        cell = ws.cell(
                            row=row_idx, column=col_idx, value=value)
                        if cbg_link:
                            cell.hyperlink = cbg_link
                            cell.font = Font(
                                color="0000FF", underline="single")
                        else:
                            cell.font = Font()
                    else:
                        # 数值格式化
                        if isinstance(value, (int, float)):
                            if col_name in ['price', 'improved_value'] or col_name in value_breakdown_columns:
                                # 价格显示
                                display_value = f"{value:.1f}" if value > 0 else "0"
                            elif '_ratio' in col_name:
                                # 比率显示
                                display_value = f"{value:.1f}%" if value != 0 else "0%"
                            else:
                                display_value = f"{value:.2f}" if isinstance(
                                    value, float) else str(value)
                        else:
                            display_value = str(
                                value) if value is not None else ""

                        cell = ws.cell(
                            row=row_idx, column=col_idx, value=display_value)

                    # 设置单元格样式
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center")
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )

                    # 特殊颜色
                    if col_name == 'price':
                        cell.fill = PatternFill(
                            start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                    elif col_name == 'improved_value':
                        cell.fill = PatternFill(
                            start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                    elif col_name == 'improved_diff_ratio':
                        # 根据差价比率设置颜色
                        value = float(data.get(col_name, 0))
                        if value > 20:  # 高溢价
                            cell.fill = PatternFill(
                                start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                        elif value < -20:  # 高折价
                            cell.fill = PatternFill(
                                start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                    elif col_name in value_breakdown_columns:
                        # 价值分解列使用浅蓝色
                        cell.fill = PatternFill(
                            start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")

            # 调整列宽
            for col_idx, col_name in enumerate(final_columns, 1):
                if col_name in ['equip_id']:
                    ws.column_dimensions[openpyxl.utils.get_column_letter(
                        col_idx)].width = 25
                elif col_name in ['price', 'improved_value'] or col_name in value_breakdown_columns:
                    ws.column_dimensions[openpyxl.utils.get_column_letter(
                        col_idx)].width = 15
                elif '_ratio' in col_name:
                    ws.column_dimensions[openpyxl.utils.get_column_letter(
                        col_idx)].width = 18
                else:
                    ws.column_dimensions[openpyxl.utils.get_column_letter(
                        col_idx)].width = 12

            # 冻结首行
            ws.freeze_panes = "A2"

            # 添加统计工作表
            stats_ws = wb.create_sheet("统计汇总")
            
            # 计算统计数据
            improved_values = [d['improved_value'] for d in export_data]
            prices = [d['price'] for d in export_data]
            
            improved_avg = sum(improved_values) / len(improved_values)
            price_avg = sum(prices) / len(prices)
            
            # 计算准确率（差价比率在±20%内的比例）
            improved_accurate = sum(1 for d in export_data if abs(d['improved_diff_ratio']) <= 20) / len(export_data) * 100
            
            stats_data = [
                ["统计项目", "数值", "说明"],
                ["总角色数", len(export_data), ""],
                ["平均估值", f"{improved_avg:.1f}", "改进算法平均估值"],
                ["平均实际价格", f"{price_avg:.1f}", "市场平均价格"],
                ["估值准确率(±20%)", f"{improved_accurate:.1f}%", "差价比率在±20%内的比例"],
                ["价值分解项数", len(value_breakdown_columns), "改进算法包含的价值分解维度"],
                ["折价系数应用", "是", "是否应用市场折价系数"],
            ]

            for row_idx, row_data in enumerate(stats_data, 1):
                for col_idx, value in enumerate(row_data, 1):
                    cell = stats_ws.cell(
                        row=row_idx, column=col_idx, value=value)
                    if row_idx == 1:  # 表头
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(
                            start_color="366092", end_color="366092", fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )

            # 调整统计表列宽
            for col_idx in range(1, 4):
                stats_ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

            # 保存文件
            wb.save(output_path)
            print(f"Excel文件已保存到: {output_path}")
            print(f"包含 {len(export_data)} 条角色数据")
            print(f"价值分解列: {value_breakdown_columns}")
            print(f"改进算法准确率: {improved_accurate:.1f}%")

            return output_path

        except ImportError:
            print("需要安装openpyxl库: pip install openpyxl")
            return None
        except Exception as e:
            print(f"导出Excel文件失败: {e}")
            import traceback
            traceback.print_exc()
            return None


if __name__ == "__main__":
    try:
        # 自动查找/data/文件夹下的所有.db文件
        db_files = glob.glob("data/*.db")
        if not db_files:
            print("未找到数据库文件，请确保/data/文件夹下有.db文件")
            exit(1)
        # 选择第一个.db文件
        db_path = db_files[0]
        print(f"使用数据库文件: {db_path}")

        evaluator = RuleEvaluator()
        print("开始加载数据...")
        # 加载数据
        evaluator.batch_evaluate(db_path)
        print("数据加载完成")

        print("\n=== 导出改进算法结果 ===")
        # 导出改进算法结果到Excel
        comparison_output_path = evaluator.export_comparison_to_excel(
            './output/results_comparison.xlsx')
        if comparison_output_path:
            print(f"改进算法结果已成功导出到: {comparison_output_path}")
        else:
            print("导出改进算法结果Excel失败")

    except Exception as e:
        print(f"程序运行出错: {str(e)}")
        import traceback
        print(traceback.format_exc())
