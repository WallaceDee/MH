import numpy as np
import pandas as pd
from typing import Dict, Any, List, Optional, Union, Tuple
import logging
import json
from datetime import datetime
import warnings
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

try:
    from .market_data_collector import MarketDataCollector
    from .character_evaluator import CharacterEvaluator
except ImportError:
    from market_data_collector import MarketDataCollector
    from character_evaluator import CharacterEvaluator

warnings.filterwarnings('ignore')


class MarketAnchorValuator:
    """市场锚定估价器 - 基于市场相似角色的价格锚定估价"""
    
    def __init__(self, market_data_collector: Optional[MarketDataCollector] = None):
        """
        初始化市场锚定估价器
        
        Args:
            market_data_collector: 市场数据采集器实例，如果为None则创建新实例
        """
        self.logger = logging.getLogger(__name__)
        
        # 初始化市场数据采集器
        if market_data_collector is None:
            self.market_collector = MarketDataCollector()
        else:
            self.market_collector = market_data_collector
        
        # 初始化规则估价器作为保底
        self.rule_evaluator = CharacterEvaluator()
        
        # 定义特征容忍度配置
        self.feature_tolerances = {
            'level': 0.05,           # 等级容忍度5%
            'all_new_point': 0.0,    # 乾元丹必须完全一致
            'school_history_count': 0.5,  # 历史门派数容忍度50%
            'packet_page': 1.0,      # 行囊页数容忍度100%
            'allow_pet_count': 1.0,  # 召唤兽上限容忍度100%
            
            # 修炼系统特征 - 中等容忍度
            'total_cultivation': 0.15,  # 修炼容忍度15%
            'total_beast_cultivation': 0.20,  # 召唤兽修炼容忍度20%
            'expt_ski1': 0.15,       # 攻击修炼
            'expt_ski2': 0.15,       # 防御修炼
            'expt_ski3': 0.15,       # 法术修炼
            'expt_ski4': 0.15,       # 抗法修炼
            'expt_ski5': 0.15,       # 猎术修炼
            'beast_ski1': 0.20,      # 召唤兽攻击修炼
            'beast_ski2': 0.20,      # 召唤兽防御修炼
            'beast_ski3': 0.20,      # 召唤兽法术修炼
            'beast_ski4': 0.20,      # 召唤兽抗法修炼
            
            # 经验和成长系统 - 较宽松
            'sum_exp': 0.30,         # 总经验容忍度30%
            'three_fly_lv': 0.0,     # 化圣等级必须一致
            'yushoushu_skill': 1.0,  # 育兽术等级容忍度100%
            
            # 其他增值系统
            'lingyou_count': 0.50,   # 灵佑次数容忍度50%
            'jiyuan_amount': 0.50,   # 机缘值容忍度50%
            'xianyu_amount': 0.80,   # 仙玉容忍度80%
            'learn_cash': 0.80,      # 储备金容忍度80%
            
            # 宠物和法宝系统
            'premium_pet_count': 0.20,   # 特殊宠物容忍度20%
            'premium_fabao_count': 0.30,  # 法宝容忍度30%
            'hight_grow_rider_count': 0.20,  # 高成长坐骑容忍度20%
            
            # 技能系统 - 组合特征
            'total_school_skills': 0.15,    # 师门技能总和
            'avg_school_skills': 0.10,      # 师门技能平均值
            'total_life_skills': 0.5,       # 生活技能总和
            'high_life_skills_count': 0.30, # 高等级生活技能数量
            'total_qiangzhuang_shensu': 0.20, # 强壮神速总和
            
            # 特殊状态
            'qianyuandan_breakthrough': 0.0,  # 乾元丹突破必须一致
        }
        
        # 定义关键特征权重（用于相似度计算）
        self.feature_weights = {
            # 最重要特征 - 核心影响价值
            'level': 1.0,
            'all_new_point': 1.0,
            'qianyuandan_breakthrough': 1.0,
            
            # 很重要特征 - 显著影响价值
            'total_cultivation': 0.9,
            'school_history_count': 0.8,
            
            # 重要特征 - 中等影响
            'total_beast_cultivation': 0.7,
            'premium_fabao_count': 0.6,
            'packet_page': 0.6,
            'avg_school_skills': 0.6,
            'three_fly_lv': 0.6,
            
            # 修炼细节特征 - 中等权重
            'expt_ski1': 0.5, 'expt_ski2': 0.5, 'expt_ski3': 0.5, 'expt_ski4': 0.5, 'expt_ski5': 0.5,
            'beast_ski1': 0.4, 'beast_ski2': 0.4, 'beast_ski3': 0.4, 'beast_ski4': 0.4,
            
            # 辅助特征 - 较低权重
            'sum_exp': 0.4,
            'lingyou_count': 0.4,
            'allow_pet_count': 0.4,
            'yushoushu_skill': 0.4,
            'total_school_skills': 0.4,
            'premium_pet_count': 0.4,
            'hight_grow_rider_count': 0.4,
            
            # 次要特征 - 低权重
            'jiyuan_amount': 0.3,
            'total_life_skills': 0.3,
            'high_life_skills_count': 0.3,
            'total_qiangzhuang_shensu': 0.3,
            
            # 可选特征 - 最低权重
            'xianyu_amount': 0.2,
            'learn_cash': 0.2,
        }
        
        print("市场锚定估价器初始化完成")
    
    def find_market_anchors(self, 
                           target_features: Dict[str, Any],
                           similarity_threshold: float = 0.6,
                           max_anchors: int = 30) -> List[Dict[str, Any]]:
        """
        寻找市场锚点角色
        
        Args:
            target_features: 目标角色特征字典
            similarity_threshold: 相似度阈值（0-1）
            max_anchors: 最大锚点数量
            
        Returns:
            List[Dict[str, Any]]: 锚点角色列表，每个元素包含：
                - similarity: 相似度分数
                - price: 价格
                - equip_id: 装备ID
                - features: 完整特征数据
        """
        try:
            print(f"开始寻找市场锚点，相似度阈值: {similarity_threshold}")
            
            # 获取市场数据
            market_data = self.market_collector.get_market_data()
            
            if market_data.empty:
                print("市场数据为空，无法找到锚点")
                return []
            
            # 创建目标特征向量
            target_vector = self._create_feature_vector(target_features)
            
            # 计算所有市场角色的相似度
            anchor_candidates = []
            
            for equip_id, market_row in market_data.iterrows():
                try:
                    # 计算相似度
                    similarity = self._calculate_similarity(target_features, market_row.to_dict())
                    
                    if similarity >= similarity_threshold:
                        anchor_candidates.append({
                            'equip_id': equip_id,
                            'similarity': similarity,
                            'price': market_row.get('price', 0),
                            'features': market_row.to_dict()
                        })
                        
                except Exception as e:
                    self.logger.warning(f"计算角色 {equip_id} 相似度时出错: {e}")
                    continue
            
            # 按相似度排序
            anchor_candidates.sort(key=lambda x: x['similarity'], reverse=True)
            
            # 返回前N个锚点
            anchors = anchor_candidates[:max_anchors]
            
            print(f"找到 {len(anchors)} 个市场锚点角色")
            if anchors:
                print(f"相似度范围: {anchors[-1]['similarity']:.3f} - {anchors[0]['similarity']:.3f}")
                print(f"价格范围: {min(a['price'] for a in anchors):.1f} - {max(a['price'] for a in anchors):.1f}")
            
            return anchors
            
        except Exception as e:
            self.logger.error(f"寻找市场锚点失败: {e}")
            return []
    
    def _create_feature_vector(self, features: Dict[str, Any]) -> np.ndarray:
        """
        创建特征向量
        
        Args:
            features: 特征字典
            
        Returns:
            np.ndarray: 标准化的特征向量
        """
        vector = []
        for feature_name in self.feature_tolerances.keys():
            value = features.get(feature_name, 0)
            # 简单标准化
            if feature_name == 'level':
                normalized_value = value / 200.0  # 假设最高等级200
            elif feature_name in ['sum_exp']:
                normalized_value = min(value / 10000.0, 1.0)  # 经验标准化
            elif feature_name in ['total_cultivation', 'total_beast_cultivation']:
                normalized_value = value / 400.0  # 修炼标准化（假设最高100*4）
            elif feature_name in ['total_gem_level']:
                normalized_value = value / 1000.0  # 宝石标准化
            else:
                normalized_value = min(value / 50.0, 1.0)  # 通用标准化
            
            vector.append(normalized_value)
        
        return np.array(vector)
    
    def _calculate_similarity(self, 
                            target_features: Dict[str, Any], 
                            market_features: Dict[str, Any]) -> float:
        """
        计算两个角色特征的相似度
        
        Args:
            target_features: 目标角色特征
            market_features: 市场角色特征
            
        Returns:
            float: 相似度分数（0-1）
        """
        total_weight = 0
        weighted_similarity = 0
        epsilon = 1e-8  # 防止除零的小值
        
        for feature_name, tolerance in self.feature_tolerances.items():
            if feature_name not in target_features and feature_name not in market_features:
                continue
                
            target_val = target_features.get(feature_name, 0)
            market_val = market_features.get(feature_name, 0)
            weight = self.feature_weights.get(feature_name, 0.5)
            
            # 计算特征差异百分比
            if target_val == 0 and market_val == 0:
                # 两者都为0，完全匹配
                feature_similarity = 1.0
            elif target_val == 0 or market_val == 0:
                # 一个为0一个不为0，根据特征类型决定相似度
                if feature_name in ['all_new_point', 'qianyuandan_breakthrough', 'three_fly_lv']:
                    # 关键特征必须一致
                    feature_similarity = 0.0
                else:
                    # 非关键特征给予部分相似度
                    feature_similarity = 0.3
            else:
                # 计算相对差异，添加epsilon防止除零
                denominator = max(abs(target_val), epsilon)
                diff_ratio = abs(target_val - market_val) / denominator
                
                if diff_ratio <= tolerance:
                    # 在容忍度内，相似度为1
                    feature_similarity = 1.0
                elif diff_ratio <= tolerance * 2:
                    # 超出容忍度但在2倍范围内，线性递减
                    feature_similarity = max(0, 1.0 - (diff_ratio - tolerance) / max(tolerance, epsilon))
                else:
                    # 差异太大，相似度为0
                    feature_similarity = 0.0
            
            weighted_similarity += feature_similarity * weight
            total_weight += weight
        
        # 计算加权平均相似度
        if total_weight > 0:
            final_similarity = weighted_similarity / total_weight
        else:
            final_similarity = 0.0
        
        return final_similarity
    
    def calculate_value(self, 
                       target_features: Dict[str, Any],
                       strategy: str = 'fair_value') -> Dict[str, Any]:
        """
        计算角色价值
        
        Args:
            target_features: 目标角色特征字典
            strategy: 定价策略 ('competitive', 'fair_value', 'premium')
            
        Returns:
            Dict[str, Any]: 估价结果，包含：
                - estimated_price: 估算价格
                - anchor_count: 锚点数量
                - price_range: 价格范围
                - confidence: 置信度
                - strategy_used: 使用的策略
                - fallback_used: 是否使用了保底估价
        """
        try:
            print(f"开始计算角色价值，策略: {strategy}")
            
            # 寻找市场锚点
            anchors = self.find_market_anchors(target_features)
            
            if len(anchors) == 0:
                print("未找到市场锚点，使用规则估价保底")
                return self._fallback_valuation(target_features, strategy)
            
            # 提取锚点价格
            anchor_prices = [anchor['price'] for anchor in anchors]
            anchor_similarities = [anchor['similarity'] for anchor in anchors]
            
            # 根据策略计算估价
            if strategy == 'competitive':
                # 竞争性定价：25%分位数 × 0.95
                estimated_price = np.percentile(anchor_prices, 25) * 0.95
            elif strategy == 'premium':
                # 溢价定价：75%分位数 × 1.05
                estimated_price = np.percentile(anchor_prices, 75) * 1.05
            else:  # fair_value
                # 公允价值：相似度加权中位数
                estimated_price = self._weighted_median(anchor_prices, anchor_similarities)
            
            # 计算置信度
            confidence = self._calculate_confidence(anchors, len(anchor_prices))
            
            # 构建结果
            result = {
                'estimated_price': round(estimated_price, 1),
                'anchor_count': len(anchors),
                'price_range': {
                    'min': min(anchor_prices),
                    'max': max(anchor_prices),
                    'mean': np.mean(anchor_prices),
                    'median': np.median(anchor_prices)
                },
                'confidence': confidence,
                'strategy_used': strategy,
                'fallback_used': False,
                'anchors': anchors[:5]  # 返回前5个最相似的锚点用于展示
            }
            
            print(f"估价完成: {estimated_price:.1f}，基于 {len(anchors)} 个锚点，置信度: {confidence:.2f}")
            
            return result
            
        except Exception as e:
            self.logger.error(f"计算角色价值失败: {e}")
            return self._fallback_valuation(target_features, strategy)
    
    def _weighted_median(self, values: List[float], weights: List[float]) -> float:
        """
        计算加权中位数
        
        Args:
            values: 数值列表
            weights: 权重列表
            
        Returns:
            float: 加权中位数
        """
        if not values or not weights:
            return 0.0
        
        # 将数值和权重配对并排序
        paired = list(zip(values, weights))
        paired.sort(key=lambda x: x[0])
        
        # 计算累积权重
        total_weight = sum(weights)
        cumulative_weight = 0
        
        for value, weight in paired:
            cumulative_weight += weight
            if cumulative_weight >= total_weight * 0.5:
                return value
        
        return paired[-1][0]  # fallback
    
    def _calculate_confidence(self, anchors: List[Dict[str, Any]], anchor_count: int) -> float:
        """
        计算估价置信度
        
        Args:
            anchors: 锚点列表
            anchor_count: 锚点数量
            
        Returns:
            float: 置信度（0-1）
        """
        # 基础置信度基于锚点数量
        base_confidence = min(anchor_count / 20.0, 1.0)  # 20个锚点达到满分
        
        # 相似度加成
        if anchors:
            avg_similarity = np.mean([anchor['similarity'] for anchor in anchors])
            similarity_bonus = avg_similarity * 0.3
        else:
            similarity_bonus = 0
        
        # 价格稳定性加成
        if len(anchors) >= 3:
            prices = [anchor['price'] for anchor in anchors]
            price_cv = np.std(prices) / np.mean(prices) if np.mean(prices) > 0 else 1.0
            stability_bonus = max(0, (0.5 - price_cv) * 0.4)  # CV低于0.5时有加成
        else:
            stability_bonus = 0
        
        final_confidence = min(base_confidence + similarity_bonus + stability_bonus, 1.0)
        return final_confidence
    
    def _fallback_valuation(self, 
                           target_features: Dict[str, Any], 
                           strategy: str) -> Dict[str, Any]:
        """
        保底估价（使用规则估价器）
        
        Args:
            target_features: 目标角色特征
            strategy: 定价策略
            
        Returns:
            Dict[str, Any]: 保底估价结果
        """
        try:
            # 使用规则估价器计算基础价值
            # 构造简单的character_data
            character_data = {
                'level': target_features.get('level', 109),
                'all_new_point': target_features.get('all_new_point', 0),
                'sum_exp': target_features.get('sum_exp', 0),
                'school_skills': target_features.get('school_skills', []),
                'life_skills': target_features.get('life_skills', []),
                # 添加其他必要字段...
            }
            
            # 使用规则估价器
            rule_result = self.rule_evaluator.evaluate(character_data, use_improved=True)
            base_value = rule_result.get('base_value', 3000)  # 默认保底3000
            
            # 根据策略调整
            if strategy == 'competitive':
                estimated_price = base_value * 0.9
            elif strategy == 'premium':
                estimated_price = base_value * 1.2
            else:  # fair_value
                estimated_price = base_value
            
            return {
                'estimated_price': round(estimated_price, 1),
                'anchor_count': 0,
                'price_range': {
                    'min': estimated_price,
                    'max': estimated_price,
                    'mean': estimated_price,
                    'median': estimated_price
                },
                'confidence': 0.3,  # 保底估价置信度较低
                'strategy_used': f"{strategy}_fallback",
                'fallback_used': True,
                'anchors': []
            }
            
        except Exception as e:
            self.logger.error(f"保底估价失败: {e}")
            # 最终保底价格
            fallback_price = 2000.0
            return {
                'estimated_price': fallback_price,
                'anchor_count': 0,
                'price_range': {'min': fallback_price, 'max': fallback_price, 'mean': fallback_price, 'median': fallback_price},
                'confidence': 0.1,
                'strategy_used': 'emergency_fallback',
                'fallback_used': True,
                'anchors': []
            }
    
    def value_distribution_report(self, target_features: Dict[str, Any]) -> Dict[str, Any]:
        """
        生成价值分布报告
        
        Args:
            target_features: 目标角色特征字典
            
        Returns:
            Dict[str, Any]: 价值分布报告，包含：
                - min_price, max_price, median_price: 价格统计
                - percentile_25, percentile_75: 分位数
                - recommended_competitive, recommended_fair: 推荐价格
                - anchor_count: 锚点数量
                - price_distribution: 价格分布直方图数据
        """
        try:
            print("生成价值分布报告...")
            
            # 寻找锚点
            anchors = self.find_market_anchors(target_features, similarity_threshold=0.5, max_anchors=50)
            
            if len(anchors) == 0:
                return {
                    'error': '未找到足够的市场锚点',
                    'min_price': 0,
                    'max_price': 0,
                    'median_price': 0,
                    'anchor_count': 0
                }
            
            # 提取价格数据
            prices = [anchor['price'] for anchor in anchors]
            similarities = [anchor['similarity'] for anchor in anchors]
            
            # 计算统计量
            min_price = min(prices)
            max_price = max(prices)
            median_price = np.median(prices)
            percentile_25 = np.percentile(prices, 25)
            percentile_75 = np.percentile(prices, 75)
            
            # 计算推荐价格
            competitive_result = self.calculate_value(target_features, 'competitive')
            fair_result = self.calculate_value(target_features, 'fair_value')
            
            # 生成价格分布直方图数据
            hist_bins = 10
            hist_counts, hist_edges = np.histogram(prices, bins=hist_bins)
            
            # 构建报告
            report = {
                'min_price': float(min_price),
                'max_price': float(max_price),
                'median_price': float(median_price),
                'percentile_25': float(percentile_25),
                'percentile_75': float(percentile_75),
                'recommended_competitive': competitive_result['estimated_price'],
                'recommended_fair': fair_result['estimated_price'],
                'anchor_count': len(anchors),
                'price_distribution': {
                    'bins': [float(edge) for edge in hist_edges],
                    'counts': [int(count) for count in hist_counts]
                },
                'anchor_details': [
                    {
                        'equip_id': anchor['equip_id'],
                        'price': anchor['price'],
                        'similarity': round(anchor['similarity'], 3),
                        'level': anchor['features'].get('level', 0),
                        'cultivation': anchor['features'].get('total_cultivation', 0)
                    }
                    for anchor in anchors[:10]  # 返回前10个详细信息
                ]
            }
            
            print(f"价值分布报告生成完成，基于 {len(anchors)} 个锚点")
            
            return report
            
        except Exception as e:
            self.logger.error(f"生成价值分布报告失败: {e}")
            return {'error': str(e)}
    
    def batch_valuation(self, 
                       character_list: List[Dict[str, Any]],
                       strategy: str = 'fair_value') -> List[Dict[str, Any]]:
        """
        批量估价
        
        Args:
            character_list: 角色特征列表
            strategy: 定价策略
            
        Returns:
            List[Dict[str, Any]]: 批量估价结果列表
        """
        results = []
        
        print(f"开始批量估价，共 {len(character_list)} 个角色")
        
        for i, character_features in enumerate(character_list):
            try:
                result = self.calculate_value(character_features, strategy)
                result['character_index'] = i
                results.append(result)
                
                if (i + 1) % 10 == 0:
                    print(f"已完成 {i + 1}/{len(character_list)} 个角色的估价")
                    
            except Exception as e:
                self.logger.error(f"批量估价第 {i+1} 个角色失败: {e}")
                results.append({
                    'character_index': i,
                    'estimated_price': 0,
                    'error': str(e)
                })
        
        print(f"批量估价完成，成功估价 {len([r for r in results if 'error' not in r])} 个角色")
        
        return results
    
    def export_valuation_report_to_excel(self, 
                                       target_features: Dict[str, Any],
                                       output_file: str = None,
                                       include_anchors: bool = True) -> str:
        """
        导出估价报告到Excel文件
        
        Args:
            target_features: 目标角色特征字典
            output_file: 输出文件路径，如果为None则自动生成
            include_anchors: 是否包含锚点角色详细信息
            
        Returns:
            str: 导出的文件路径
        """
        try:
            print("开始生成Excel估价报告...")
            
            # 生成文件名
            if output_file is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_file = f"角色估价报告_{timestamp}.xlsx"
            
            # 获取估价结果和分布报告
            fair_result = self.calculate_value(target_features, 'fair_value')
            competitive_result = self.calculate_value(target_features, 'competitive')
            premium_result = self.calculate_value(target_features, 'premium')
            distribution_report = self.value_distribution_report(target_features)
            
            # 创建Excel工作簿
            wb = Workbook()
            
            # 删除默认工作表
            wb.remove(wb.active)
            
            # 1. 创建估价摘要工作表
            self._create_summary_sheet(wb, target_features, fair_result, competitive_result, premium_result)
            
            # 2. 创建价格分布工作表
            self._create_distribution_sheet(wb, distribution_report)
            
            # 3. 创建相似角色列表工作表
            if include_anchors and fair_result.get('anchor_count', 0) > 0:
                self._create_anchors_sheet(wb, target_features)
            
            # 保存文件
            wb.save(output_file)
            print(f"Excel报告已保存至: {output_file}")
            
            return output_file
            
        except Exception as e:
            self.logger.error(f"导出Excel报告失败: {e}")
            raise e
    
    def _create_summary_sheet(self, wb: Workbook, 
                            target_features: Dict[str, Any],
                            fair_result: Dict[str, Any],
                            competitive_result: Dict[str, Any],
                            premium_result: Dict[str, Any]):
        """创建估价摘要工作表"""
        ws = wb.create_sheet("估价摘要", 0)
        
        # 设置标题样式
        title_font = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        header_font = Font(name='微软雅黑', size=12, bold=True)
        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        
        normal_font = Font(name='微软雅黑', size=11)
        
        # 标题
        ws['A1'] = '梦幻西游角色估价报告'
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws.merge_cells('A1:D1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 角色基本信息
        row = 3
        ws[f'A{row}'] = '角色基本信息'
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        key_features = ['level', 'all_new_point', 'total_cultivation', 'total_beast_cultivation', 
                       'school_history_count', 'premium_fabao_count', 'sum_exp']
        
        for feature in key_features:
            if feature in target_features:
                ws[f'A{row}'] = self._get_feature_display_name(feature)
                ws[f'B{row}'] = target_features[feature]
                ws[f'A{row}'].font = normal_font
                ws[f'B{row}'].font = normal_font
                row += 1
        
        # 估价结果
        row += 1
        ws[f'A{row}'] = '估价结果'
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        pricing_data = [
            ('竞争性定价', competitive_result['estimated_price'], competitive_result['confidence']),
            ('公允价值定价', fair_result['estimated_price'], fair_result['confidence']),
            ('溢价定价', premium_result['estimated_price'], premium_result['confidence']),
        ]
        
        ws[f'A{row}'] = '定价策略'
        ws[f'B{row}'] = '估算价格(万)'
        ws[f'C{row}'] = '置信度'
        ws[f'D{row}'] = '锚点数量'
        
        for i in range(4):
            ws[f'{chr(65+i)}{row}'].font = header_font
            ws[f'{chr(65+i)}{row}'].fill = header_fill
        
        row += 1
        for strategy, price, confidence in pricing_data:
            ws[f'A{row}'] = strategy
            ws[f'B{row}'] = f"{price:.1f}"
            ws[f'C{row}'] = f"{confidence:.2%}"
            ws[f'D{row}'] = fair_result['anchor_count']
            
            for i in range(4):
                ws[f'{chr(65+i)}{row}'].font = normal_font
            row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
    
    def _create_distribution_sheet(self, wb: Workbook, distribution_report: Dict[str, Any]):
        """创建价格分布工作表"""
        ws = wb.create_sheet("价格分布分析")
        
        if 'error' in distribution_report:
            ws['A1'] = f"价格分布分析失败: {distribution_report['error']}"
            return
        
        # 标题样式
        header_font = Font(name='微软雅黑', size=12, bold=True)
        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        normal_font = Font(name='微软雅黑', size=11)
        
        # 价格统计
        ws['A1'] = '价格统计信息'
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws.merge_cells('A1:B1')
        
        stats_data = [
            ('最低价格', f"{distribution_report['min_price']:.1f}万"),
            ('最高价格', f"{distribution_report['max_price']:.1f}万"),
            ('中位价格', f"{distribution_report['median_price']:.1f}万"),
            ('25%分位数', f"{distribution_report['percentile_25']:.1f}万"),
            ('75%分位数', f"{distribution_report['percentile_75']:.1f}万"),
            ('锚点数量', f"{distribution_report['anchor_count']}个"),
        ]
        
        row = 2
        for label, value in stats_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = normal_font
            ws[f'B{row}'].font = normal_font
            row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
    
    def _create_anchors_sheet(self, wb: Workbook, target_features: Dict[str, Any]):
        """创建相似角色列表工作表"""
        ws = wb.create_sheet("相似角色列表")
        
        # 获取锚点数据
        anchors = self.find_market_anchors(target_features, similarity_threshold=0.4, max_anchors=100)
        
        if not anchors:
            ws['A1'] = "未找到相似角色"
            return
        
        # 标题样式
        header_font = Font(name='微软雅黑', size=12, bold=True)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font.color = 'FFFFFF'
        
        normal_font = Font(name='微软雅黑', size=10)
        
        # 设置表头
        headers = [
            '序号', '角色ID', '价格(万)', '相似度', '等级', '潜能点', '总修炼', 
            '召唤兽修炼', '门派数', '法宝数量', '藏宝阁链接'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 填充数据
        for idx, anchor in enumerate(anchors, 1):
            row = idx + 1
            features = anchor['features']
            
            # 基本信息
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=str(anchor['equip_id']))
            ws.cell(row=row, column=3, value=f"{anchor['price']:.1f}")
            ws.cell(row=row, column=4, value=f"{anchor['similarity']:.3f}")
            ws.cell(row=row, column=5, value=features.get('level', 0))
            ws.cell(row=row, column=6, value=features.get('all_new_point', 0))
            ws.cell(row=row, column=7, value=features.get('total_cultivation', 0))
            ws.cell(row=row, column=8, value=features.get('total_beast_cultivation', 0))
            ws.cell(row=row, column=9, value=features.get('school_history_count', 0))
            ws.cell(row=row, column=10, value=features.get('premium_fabao_count', 0))
            
            # 藏宝阁链接
            cbg_url = f"https://xyq.cbg.163.com/cgi/mweb/equip/equip_detail?equip_id={anchor['equip_id']}"
            ws.cell(row=row, column=11, value=cbg_url)
            
            # 设置超链接
            from openpyxl.worksheet.hyperlink import Hyperlink
            ws.cell(row=row, column=11).hyperlink = cbg_url
            ws.cell(row=row, column=11).style = "Hyperlink"
            
            # 设置字体
            for col in range(1, 12):
                ws.cell(row=row, column=col).font = normal_font
        
        # 调整列宽
        column_widths = [6, 15, 10, 8, 6, 8, 8, 10, 8, 8, 25]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64+col)].width = width
        
        # 设置筛选
        ws.auto_filter.ref = f"A1:K{len(anchors)+1}"
        
        print(f"已导出 {len(anchors)} 个相似角色到Excel")
    
    def _get_feature_display_name(self, feature_name: str) -> str:
        """获取特征的显示名称"""
        display_names = {
            'level': '角色等级',
            'all_new_point': '潜能点数',
            'total_cultivation': '总修炼等级',
            'total_beast_cultivation': '召唤兽总修炼',
            'school_history_count': '历史门派数',
            'premium_fabao_count': '法宝数量',
            'sum_exp': '总经验',
            'three_fly_lv': '化圣等级',
            'premium_pet_count': '特殊宠物数',
            'hight_grow_rider_count': '高成长坐骑数',
            'packet_page': '行囊页数',
            'allow_pet_count': '召唤兽上限',
        }
        return display_names.get(feature_name, feature_name)


if __name__ == "__main__":
    # 测试代码
    try:
        # 初始化估价器
        valuator = MarketAnchorValuator()
        
        # 构造测试角色特征
        test_features = {
            'level': 129,
            'total_cultivation': 63,
            'total_beast_cultivation': 80,
            'all_new_point': 7,
            'sum_exp': 200,
            'school_history_count': 1,
            'premium_fabao_count': 2,
            'hight_grow_rider_count': 3,
            'expt_ski2': 21,       # 防御修炼
            'expt_ski3': 21,       # 法术修炼
            'expt_ski4': 21,       # 抗法修炼
            'expt_ski5': 21,       # 猎术修炼
            'beast_ski1': 20,      # 召唤兽攻击修炼
            'beast_ski2': 20,      # 召唤兽防御修炼
            'beast_ski3': 20,      # 召唤兽法术修炼
            'beast_ski4': 20,      # 召唤兽抗法修炼
        }
        
        print("=== 测试角色特征 ===")
        for key, value in test_features.items():
            print(f"{key}: {value}")
        
        # 测试估价
        print("\n=== 开始估价测试 ===")
        
        # 公允价值估价
        fair_result = valuator.calculate_value(test_features, 'fair_value')
        print(f"\n公允价值估价: {fair_result['estimated_price']:.1f}")
        print(f"锚点数量: {fair_result['anchor_count']}")
        print(f"置信度: {fair_result['confidence']:.2f}")
        
        # 竞争性估价
        competitive_result = valuator.calculate_value(test_features, 'competitive')
        print(f"\n竞争性估价: {competitive_result['estimated_price']:.1f}")
        
        # 溢价估价
        premium_result = valuator.calculate_value(test_features, 'premium')
        print(f"\n溢价估价: {premium_result['estimated_price']:.1f}")
        
        # 生成价值分布报告
        print("\n=== 生成价值分布报告 ===")
        report = valuator.value_distribution_report(test_features)
        
        if 'error' not in report:
            print(f"价格范围: {report['min_price']:.1f} - {report['max_price']:.1f}")
            print(f"中位价格: {report['median_price']:.1f}")
            print(f"推荐竞争价格: {report['recommended_competitive']:.1f}")
            print(f"推荐公允价格: {report['recommended_fair']:.1f}")
            print(f"锚点数量: {report['anchor_count']}")
        else:
            print(f"报告生成失败: {report['error']}")
        
        # 测试Excel导出
        print("\n=== 测试Excel导出 ===")
        excel_file = valuator.export_valuation_report_to_excel(test_features)
        print(f"Excel报告已生成: {excel_file}")
        
    except Exception as e:
        print(f"测试失败: {e}")
        import traceback
        traceback.print_exc() 