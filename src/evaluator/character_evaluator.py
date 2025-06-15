try:
    from .feature_extractor import FeatureExtractor
except ImportError:
    from feature_extractor import FeatureExtractor
import os
import sys
import sqlite3
import json
import logging
import numpy as np
from datetime import datetime
from typing import Dict, Any, List, Optional, Union, Tuple
import glob
import pandas as pd
import joblib

# 添加src目录到Python路径
current_dir = os.path.dirname(os.path.abspath(__file__))
src_dir = os.path.dirname(current_dir)
if src_dir not in sys.path:
    sys.path.insert(0, src_dir)


class CharacterEvaluator:
    def __init__(self, model_dir='models'):
        """
        model_dir: str, 模型保存目录
        """
        self.logger = logging.getLogger(__name__)
        self.model_dir = model_dir
        self.feature_extractor = FeatureExtractor()
        self.df = None

        # 获取当前文件所在目录，然后构建rule_setting.json的正确路径
        rule_setting_path = os.path.join(current_dir, 'rule_setting.json')
        self.rule_setting = json.load(
            open(rule_setting_path, 'r', encoding='utf-8'))


        # 装备评分系数
        self.EQUIP_COEFFS = {
            'total_gem_level': 80,       # 每级宝石价值
            'premium_skill_count': 500,  # 每个特技价值
            'set_bonus_count': 300       # 每件套装价值
        }

        # 宝宝评分系数
        self.PET_COEFFS = {
            'premium_pet_count': 1500,   # 每只特殊宝宝
            'total_pet_score': 0.8,      # 宝宝总分系数
            'max_pet_score': 1.2         # 最高宝宝分系数
        }

        # 确保模型目录存在
        if not os.path.exists(model_dir):
            os.makedirs(model_dir)

    def _connect_db(self):
        """连接数据库"""
        return sqlite3.connect(self.db_path)

    def _extract_features(self, character_data, conn=None):
        """提取角色特征

        使用 FeatureExtractor 提取特征
        """
        try:
            # 使用 FeatureExtractor 提取特征
            features = self.feature_extractor.extract_features(character_data)
            return features

        except Exception as e:
            self.logger.error(f"特征提取失败: {e}")
            raise

    def calc_base_attributes_value(self, features: Dict[str, Union[int, float, str]]) -> float:
        """
        计算基础属性价值
        """
        try:
            RULE = self.rule_setting
            value = 0
            level: int = features.get('level', 0)

            # 历史门派数加成
            school_history_count = features.get('school_history_count', 0)
            value += school_history_count * \
                RULE.get('ChangeSchoolCostPerTimes', 0)
            print(
                f"历史门派数加成: {school_history_count * RULE.get('ChangeSchoolCostPerTimes', 0)}")

            # 乾元丹加成 - 添加安全检查
            all_new_point = features.get('all_new_point', 0)
            qianyuandan_dict = RULE.get('QianyuandanCount2Value', {})
            if str(all_new_point) in qianyuandan_dict:
                value += all_new_point * qianyuandan_dict[str(all_new_point)]
            print(
                f"all_new_point: {all_new_point}")    
            print(
                f"乾元丹加成: {all_new_point * qianyuandan_dict[str(all_new_point)]}")
            # 乾元丹突破加成
            if features.get('qianyuandan_breakthrough', False):
                value += RULE.get('qianyuandanBreakthroughValue', 0)
            print(f"乾元丹突破加成: {RULE.get('qianyuandanBreakthroughValue', 0)}")
            # 月饼粽子机缘
            jiyuan_amount = features.get('jiyuan_amount', 0)
            value += jiyuan_amount * RULE.get('JiyuanValuePerCount', 0)
            print(
                f"月饼粽子机缘: {jiyuan_amount * RULE.get('JiyuanValuePerCount', 0)}")
            #  总经验
            sum_exp = features.get('sum_exp', 0)
            value += sum_exp * RULE.get('ExpValuePerHundredMillion', 0)
            print(f"总经验: {sum_exp * RULE.get('ExpValuePerHundredMillion', 0)}")
            #  师门技能
            school_skills = features.get('school_skills', [])
            school_skill_dict = RULE.get('SchoolSkillGrade2Value', {})
            tempValue = 0
            for skill_grade in school_skills:
                if str(skill_grade) in school_skill_dict:
                    value += school_skill_dict[str(skill_grade)]
                    tempValue += school_skill_dict[str(skill_grade)]
            print(f"师门技能: {tempValue}")
            #  行囊
            packet_page = features.get('packet_page', 0)
            packet_dict = RULE.get('PackagePage2Value', {})
            if str(packet_page) in packet_dict:
                value += packet_dict[str(packet_page)]
                print(f"行囊: {packet_dict[str(packet_page)]}")
            # 储备金
            learn_cash = features.get('learn_cash', 0)
            value += learn_cash/10000 * RULE.get('LearnCash2CashRate', 0)
            print(
                f"储备金: {learn_cash/10000 * RULE.get('LearnCash2CashRate', 0)}")
            # 仙玉
            xianyu_amount = features.get('xianyu_amount', 0)
            value += xianyu_amount * RULE.get('RMB2MHB', 0)
            print(f"仙玉: {xianyu_amount * RULE.get('RMB2MHB', 0)}")
            # 修炼 - 添加安全检查，移除重复计算
            max_cultivate_dict = RULE.get('MaxCultivateGrade2Value', {})
            cultivate_cost_dict = RULE.get('CultivateCostPerCount', {})
            cultivate_learn_count_dict = RULE['CultivateGrade2LearnCount']
            xiulianshangxianValue = 0
            xiulianValue = 0
            # 检查各个修炼等级上限
            for i in range(1, 6):  # 修炼1-4 (包含expt_ski5)
                max_expt_key = f'max_expt{i}'
                expt_key = f'expt_ski{i}'
                max_expt_value = features.get(max_expt_key, 0)
                expt_value = features.get(expt_key, 0)
                cost_multiplier = cultivate_cost_dict.get(str(i), 0)
                # 只计算一次，不区分是否>20
                if max_expt_value > 0 and str(max_expt_value) in max_cultivate_dict and i != 5:
                    value += max_cultivate_dict[str(max_expt_value)] * cost_multiplier
                    xiulianshangxianValue += max_cultivate_dict[str(max_expt_value)] * cost_multiplier
                    print( f"修炼上限{i}:{max_expt_value} = {max_cultivate_dict[str(max_expt_value)] * cost_multiplier}")
                value += cultivate_learn_count_dict[str(expt_value)] * cost_multiplier
                xiulianValue += cultivate_learn_count_dict[str(expt_value)] * cost_multiplier
                print( f"修炼等级{i}:{expt_value} = {cultivate_learn_count_dict[str(expt_value)] * cost_multiplier}")
            print(f"修炼等级上限: {xiulianshangxianValue}")
            print(f"修炼等级: {xiulianValue}")
            # 控制力
            control_dict = RULE.get('ControlGrade2XiulianguoCount', {})
            xiulianguo_value = RULE.get('XiulianguoValuePerCount', 0)

            bbXiulianValue = 0
            for i in range(1, 5):  # beast_ski1-4
                beast_ski_key = f'beast_ski{i}'
                beast_ski_value = features.get(beast_ski_key, 0)
                if str(beast_ski_value) in control_dict:
                    value += xiulianguo_value * control_dict[str(beast_ski_value)]
                    bbXiulianValue += xiulianguo_value * control_dict[str(beast_ski_value)]
                    print(f"控制力{i} {beast_ski_value}: {xiulianguo_value * control_dict[str(beast_ski_value)]}")
            print(f"控制力: {bbXiulianValue}")
            # 生活技能
            life_skillsValue = 0
            life_skills = features.get('life_skills', [])
            life_skill_dict = RULE.get('LifeSkillGrade2Value', {})
            for life_skill_grade in life_skills:
                if life_skill_grade >= 80 and str(life_skill_grade) in life_skill_dict:
                    value += life_skill_dict[str(life_skill_grade)]
                    life_skillsValue += life_skill_dict[str(life_skill_grade)]
            print(f"生活技能: {life_skillsValue}")
            # 强壮、神速
            qiangzhuang_shensuValue = 0
            qiangzhuang_shensu = features.get('qiangzhuang&shensu', [])
            qiangzhuang_dict = RULE.get('QiangzhuangAndShensuGrade2Value', {})
            for qiangzhuang_and_shensu_grade in qiangzhuang_shensu:
                if str(qiangzhuang_and_shensu_grade) in qiangzhuang_dict:
                    value += qiangzhuang_dict[str(qiangzhuang_and_shensu_grade)]
                    qiangzhuang_shensuValue += qiangzhuang_dict[str(
                        qiangzhuang_and_shensu_grade)]
            print(f"强壮、神速: {qiangzhuang_shensuValue}")
            
            # 灵佑次数
            lingyou_count = features.get('lingyou_count', 0)
            value += lingyou_count * RULE.get('LingyouValuePerCount', 0)
            print(f"灵佑次数：{lingyou_count}: {lingyou_count * RULE.get('LingyouValuePerCount', 0)}")

            # 高成长坐骑数量
            hight_grow_rider_count = features.get('hight_grow_rider_count', 0)
            value += hight_grow_rider_count * RULE.get('HightGrowRidePerValue', 0)
            print(f"高成长坐骑数量：{lingyou_count}: {lingyou_count * RULE.get('HightGrowRidePerValue', 0)}")

            # 召唤兽最大携带量
            allow_pet_count = features.get('allow_pet_count', 0)
            value += RULE.get('AllowPetCount2Value', {}).get(str(allow_pet_count),0)
            print(f"召唤兽最大携带量：{allow_pet_count}: {RULE.get('AllowPetCount2Value', {}).get(str(allow_pet_count),0)}")

            # 有价值法宝数量
            premium_fabao_count = features.get('premium_fabao_count', 0)
            value += premium_fabao_count * RULE.get('PremiumFabaoValuePerCount', 0)
            print(f"有价值法宝数量：{premium_fabao_count}: {premium_fabao_count * RULE.get('PremiumFabaoValuePerCount', 0)}")
            
            # 神器
            shenqi = features.get('shenqi', {})
            ShenqiAttr2Value = RULE.get('ShenqiAttr2Value', {})
            shenqiValue = 0
            if shenqi.get('same9Count', 0):
                value += shenqi['same9Count']*ShenqiAttr2Value.get('same9Count', 0)
                shenqiValue += shenqi['same9Count']*ShenqiAttr2Value.get('same9Count', 0)
            if shenqi.get('allTheSameCount', 0):
                value += shenqi['allTheSameCount']*ShenqiAttr2Value.get('allTheSameCount', 0)
                shenqiValue += shenqi['allTheSameCount']*ShenqiAttr2Value.get('allTheSameCount', 0)
            if shenqi.get('1Count', 0):
                value += shenqi['1Count']*ShenqiAttr2Value.get('1', 0)
                shenqiValue += shenqi['1Count']*ShenqiAttr2Value.get('1', 0)
            if shenqi.get('1attrCount', 0):
                value += shenqi['1attrCount']*ShenqiAttr2Value.get('1attr', 0)
                shenqiValue+= shenqi['1attrCount']*ShenqiAttr2Value.get('1attr', 0)
            if shenqi.get('2Count', 0):
                value += shenqi['2Count']*ShenqiAttr2Value.get('2', 0)
                shenqiValue += shenqi['2Count']*ShenqiAttr2Value.get('2', 0)
            if shenqi.get('2attrCount', 0):
                value += shenqi['2attrCount']*ShenqiAttr2Value.get('2attr', 0)
                shenqiValue += shenqi['2attrCount']*ShenqiAttr2Value.get('2attr', 0)
            if shenqi.get('3Count', 0):
                value += shenqi['3Count']*ShenqiAttr2Value.get('3', 0)
                shenqiValue += shenqi['3Count']*ShenqiAttr2Value.get('3', 0)
            if shenqi.get('3attrCount', 0):
                value += shenqi['3attrCount']*ShenqiAttr2Value.get('3attr', 0)
                shenqiValue += shenqi['3attrCount']*ShenqiAttr2Value.get('3attr', 0)
            
            print(f"神器:{shenqiValue}-{shenqi}")
          
            # 合计
            return value/RULE['RMB2MHB'] * RULE['MARKET_FACTOR']

        except Exception as e:
            self.logger.error(f"计算基础属性价值失败: {e}")
            import traceback
            self.logger.error(f"详细错误: {traceback.format_exc()}")
            return 0.0  # 如果计算出错，返回0

    def calc_equipment_value(self, features: Dict[str, Union[int, float, str]]) -> float:
        """
        计算装备系统价值

        Args:
            features (Dict[str, Union[int, float, str]]): 角色特征字典，包含:
                - total_gem_level (int): 宝石总等级
                - premium_skill_count (int): 特技数量
                - set_bonus_count (int): 套装数量
                - total_equip_score (float): 装备总分

        Returns:
            float: 装备系统价值
        """
        value = 0

        # 宝石价值
        value += features.get('total_gem_level', 0) * \
            self.EQUIP_COEFFS['total_gem_level']

        # 特技价值
        value += features.get('premium_skill_count', 0) * \
            self.EQUIP_COEFFS['premium_skill_count']

        # 套装价值
        value += features.get('set_bonus_count', 0) * \
            self.EQUIP_COEFFS['set_bonus_count']

        # 装备总分价值（非线性增长）
        if 'total_equip_score' in features:
            score = features['total_equip_score']
            # 指数增长：总分越高，单位分价值越高
            value += score * (1 + score / 5000)

        return value

    def calc_pets_value(self, features):
        """计算召唤兽价值"""
        value = 0

        # 特殊宝宝价值
        value += features.get('premium_pet_count', 0) * \
            self.PET_COEFFS['premium_pet_count']

        # 宝宝总分价值
        if 'total_pet_score' in features:
            value += features['total_pet_score'] * \
                self.PET_COEFFS['total_pet_score']

        # 最高宝宝价值（优质宝宝溢价）
        if 'max_pet_score' in features:
            value += features['max_pet_score'] * \
                self.PET_COEFFS['max_pet_score']

        return value

    def calc_appearance_value(self, features):
        """计算外观系统价值"""
        value = 0

        # 限量锦衣价值
        if 'limited_skin_value' in features:
            # 指数衰减模型：前几件价值高，后面边际递减
            skin_value = features['limited_skin_value']
            value += 1000 * (1 - np.exp(-skin_value / 500))

        return value

    def apply_market_factors(self, value, features):
        """应用市场因子调整"""
        # 服务器热度系数（示例）
        server_hot = {
            '万里长城': 1.15,
            '生日快乐': 1.20,
            '2008': 1.18,
            '其他': 1.0
        }
        server = features.get('server', '其他')
        server_factor = server_hot.get(server, 1.0)

        # 等级段供需系数（高等级角色溢价）
        level = features.get('level', 109)
        level_factor = 1.0 + (level - 109) * 0.02  # 109级基准

        return value * server_factor * level_factor

    def evaluate(self, character_data, use_improved=True):
        """评估角色价值
        
        Args:
            character_data: 角色数据
            use_improved: 是否使用改进的算法（默认True）
        """
        try:
            # 1.提取特征
            features = self._extract_features(character_data)
            
            if use_improved:
                # 使用改进的算法
                improved_result = self.calc_base_attributes_value_improved(features)
                
                evaluation_result = {
                    'features': features,
                    'base_value': improved_result['final_value'],
                    'value_breakdown': improved_result['value_breakdown'],
                    'total_raw_value': improved_result['total_raw_value'],
                    'algorithm_version': 'improved'
                }
            else:
                # 使用原始算法
                base_value = self.calc_base_attributes_value(features)
                
                evaluation_result = {
                    'features': features,
                    'base_value': base_value,
                    'algorithm_version': 'original'
                }

            # # 3. 装备系统修正
            # equip_value = self.calc_equipment_value(features)

            # # 4. 召唤兽修正
            # pets_value = self.calc_pets_value(features)

            # # 5. 外观修正
            # appearance_value = self.calc_appearance_value(features)

            # # 6. 汇总价值
            # total_value = base_value + equip_value + pets_value + appearance_value

            # # 7. 市场因子调整
            # final_value = self.apply_market_factors(total_value, features)

            # 构建评估结果
            # evaluation_result = {
            #     'features': features,
            #     'base_value': base_value,
            #     # 'equip_value': equip_value,
            #     # 'pets_value': pets_value,
            #     # 'appearance_value': appearance_value,
            #     # 'total_value': total_value,
            #     # 'final_value': final_value,
            #     # 'value_breakdown': {
            #     #     'base_attributes': base_value,
            #     #     'equipment': equip_value,
            #     #     'pets': pets_value,
            #     #     'appearance': appearance_value
            #     # }
            # }

            return evaluation_result

        except Exception as e:
            self.logger.error(f"角色评估失败: {e}")
            raise

    def fit(self, db_path):
        """加载数据"""
        try:
            print(f"正在连接数据库: {db_path}")
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()

            # 获取所有角色数据
            print("正在查询数据库...")
            cursor.execute("""
                SELECT 
                    c.id, c.equip_id, c.server_name, c.seller_nickname, c.level, c.price,
                    c.price_desc, c.school AS school_desc, c.area_name, c.icon_index,
                    c.kindid, c.game_ordersn, c.pass_fair_show, c.fair_show_end_time,
                    c.accept_bargain, c.status_desc, c.onsale_expire_time_desc, c.expire_time,
                    c.race, c.fly_status, c.collect_num, c.life_skills, c.school_skills,
                    c.ju_qing_skills,c.yushoushu_skill, c.all_pets_json, c.all_equip_json AS all_equip_json_desc,
                    c.all_shenqi_json, c.all_rider_json AS all_rider_json_desc, c.all_fabao_json,
                    c.ex_avt_json AS ex_avt_json_desc, c.create_time, c.update_time,
                    l.*
                FROM characters c
                LEFT JOIN large_equip_desc_data l ON c.equip_id = l.equip_id
                WHERE c.price > 0 AND l.all_equip_json =='{}' AND l.all_summon_json == '[]'
            """)

            # 获取列名
            columns = [description[0] for description in cursor.description]
            print(f"查询完成，开始获取数据...")
            rows = cursor.fetchall()
            print(f"获取到 {len(rows)} 条数据")
            # 准备训练数据
            print("开始提取特征...")
            data = []
            for i, row in enumerate(rows):
                character_data = dict(zip(columns, row))
                features = self._extract_features(character_data, conn)
                features['price'] = character_data.get('price', 0)
                features['equip_id'] = character_data.get('equip_id', '')
                data.append(features)
                if (i + 1) % 10 == 0:
                    print(f"已处理 {i + 1}/{len(rows)} 条数据")

            print("特征提取完成，转换为DataFrame...")
            # 转换为DataFrame
            self.df = pd.DataFrame(data)
            print(f"DataFrame 形状: {self.df.shape}")
            self.df.set_index('equip_id', inplace=True)

            print("数据加载和预处理完成")

        except Exception as e:
            print(f"数据加载失败: {str(e)}")
            import traceback
            print(traceback.format_exc())
            raise
        finally:
            if 'conn' in locals():
                conn.close()

    def generate_cbg_link(self, eid: str) -> Optional[str]:
        """
        生成CBG角色分享链接

        Args:
            eid (str): 装备ID，格式通常为 "服务器ID-角色ID"

        Returns:
            Optional[str]: CBG链接URL，失败时返回None

        Example:
            >>> evaluator = CharacterEvaluator()
            >>> link = evaluator.generate_cbg_link("201808081404129-230011021108")
            >>> print(link)  # "https://xyq.cbg.163.com/equip?s=230011021108&eid=201808081404129-230011021108"
        """
        if not eid:
            return None

        # 从eid中提取服务器ID
        # eid格式通常是：服务器ID-角色ID，比如 "201808081404129-230011021108"
        try:
            if '-' in str(eid):
                server_id: str = str(eid).split('-')[1]

            # 构建基础CBG链接
            base_url: str = "https://xyq.cbg.163.com/equip"
            params: str = f"s={server_id}&eid={eid}"
            link: str = f"{base_url}?{params}"

            return link
        except Exception:
            return None

    def export_to_excel(self, output_path: str = 'results.xlsx') -> Optional[str]:
        """
        将评估结果导出到Excel文件

        Args:
            output_path (str): 输出文件路径，默认为 'results.xlsx'

        Returns:
            Optional[str]: 成功时返回输出文件路径，失败时返回None

        Raises:
            ImportError: 当缺少openpyxl库时
            Exception: 当导出过程中发生错误时

        Example:
            >>> evaluator = CharacterEvaluator()
            >>> evaluator.fit('data/cbg_data_202412.db')
            >>> output_path = evaluator.export_to_excel('./output/results.xlsx')
            >>> print(f"文件已保存到: {output_path}")
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows

            if self.df is None or self.df.empty:
                print("没有数据可导出")
                return None

            print(f"开始导出 {len(self.df)} 条数据到Excel...")

            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "角色评估结果"

            # 准备数据
            export_data = []
            for equip_id, row in self.df.iterrows():
                # 获取角色基本信息
                price = row.get('price', 0)

                # 计算基准价值
                try:
                    base_value = self.calc_base_attributes_value(row.to_dict())
                except:
                    base_value = 0

                # 生成CBG链接
                cbg_link = self.generate_cbg_link(equip_id)

                # 准备一行数据
                row_data = {
                    'equip_id': equip_id,
                    'cbg_link': cbg_link,
                    'price': price,
                    'base_value': base_value,
                    'price_diff_ratio': (base_value - price) / price * 100 if price > 0 else 0,  # 计算差价比率
                    **row.to_dict()  # 包含所有features
                }
                export_data.append(row_data)

            if not export_data:
                print("没有有效数据可导出")
                return None

            # 定义列顺序：优先列 + features列
            priority_columns = ['equip_id', 'price', 'base_value', 'price_diff_ratio']

            # 获取所有features列（排除优先列）
            all_columns = list(export_data[0].keys())
            feature_columns = [
                col for col in all_columns if col not in priority_columns and col != 'cbg_link']

            # 最终列顺序
            final_columns = priority_columns + feature_columns

            # 设置表头
            headers = []
            for col in final_columns:
                if col == 'equip_id':
                    headers.append('装备ID')
                elif col == 'price':
                    headers.append('价格')
                elif col == 'base_value':
                    headers.append('基准价值')
                elif col == 'price_diff_ratio':
                    headers.append('差价比率(%)')
                else:
                    headers.append(col)

            # 写入表头
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(
                    horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

            # 写入数据
            for row_idx, data in enumerate(export_data, 2):
                for col_idx, col_name in enumerate(final_columns, 1):
                    value = data.get(col_name, '')

                    # 特殊处理角色名列（添加超链接）
                    if col_name == 'equip_id':
                        cbg_link = data.get('cbg_link')
                        cell = ws.cell(
                            row=row_idx, column=col_idx, value=value)
                        if cbg_link:
                            cell.hyperlink = cbg_link
                            cell.font = Font(
                                color="0000FF", underline="single")
                        else:
                            cell.font = Font()
                    else:
                        # 数值格式化
                        if isinstance(value, (int, float)):
                            if col_name == 'price':
                                # 价格显示为万元
                                display_value = f"{value:.1f}" if value > 0 else "0"
                            else:
                                display_value = f"{value:.2f}" if isinstance(
                                    value, float) else str(value)
                        else:
                            display_value = str(
                                value) if value is not None else ""

                        cell = ws.cell(
                            row=row_idx, column=col_idx, value=display_value)

                    # 设置单元格样式
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center")
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )

                    # 价格列特殊颜色
                    if col_name == 'price':
                        cell.fill = PatternFill(
                            start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                    elif col_name == 'base_value':
                        cell.fill = PatternFill(
                            start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                    elif col_name == 'price_diff_ratio':
                        # 根据差价比率设置颜色
                        value = float(cell.value) if cell.value != '' else 0
                        if value > 20:  # 高溢价
                            cell.fill = PatternFill(
                                start_color="FF0000", end_color="FF0000", fill_type="solid")
                            cell.font = Font(color="FFFFFF")
                        elif value < -20:  # 高折价
                            cell.fill = PatternFill(
                                start_color="00FF00", end_color="00FF00", fill_type="solid")
                            cell.font = Font(color="000000")

            # 调整列宽
            for col_idx, col_name in enumerate(final_columns, 1):
                if col_name in ['equip_id']:
                    ws.column_dimensions[openpyxl.utils.get_column_letter(
                        col_idx)].width = 25
                elif col_name in ['price', 'base_value', 'price_diff_ratio']:
                    ws.column_dimensions[openpyxl.utils.get_column_letter(
                        col_idx)].width = 15
                else:
                    ws.column_dimensions[openpyxl.utils.get_column_letter(
                        col_idx)].width = 12

            # 冻结首行
            ws.freeze_panes = "A2"

            # 添加统计信息工作表
            stats_ws = wb.create_sheet("数据统计")
            stats_data = [
                ["统计项目", "数值"],
                ["总角色数", len(export_data)],
                ["平均价格", f"{self.df['price'].mean():.2f}"],
                ["最高价格", f"{self.df['price'].max():.2f}"],
                ["最低价格", f"{self.df['price'].min():.2f}"],
                ["价格中位数", f"{self.df['price'].median():.2f}"],
            ]

            for row_idx, row_data in enumerate(stats_data, 1):
                for col_idx, value in enumerate(row_data, 1):
                    cell = stats_ws.cell(
                        row=row_idx, column=col_idx, value=value)
                    if row_idx == 1:  # 表头
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(
                            start_color="366092", end_color="366092", fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")

            # 保存文件
            wb.save(output_path)
            print(f"Excel文件已保存到: {output_path}")
            print(f"包含 {len(export_data)} 条角色数据")
            print(f"优先列: {priority_columns}")
            print(f"特征列数: {len(feature_columns)}")

            return output_path

        except ImportError:
            print("需要安装openpyxl库: pip install openpyxl")
            return None
        except Exception as e:
            print(f"导出Excel失败: {e}")
            import traceback
            traceback.print_exc()
            return None

    def calc_base_attributes_value_improved(self, features: Dict[str, Union[int, float, str]]) -> Dict[str, float]:
        """
        简化的改进基础属性价值计算，只引入市场折价系数
        
        Args:
            features: 角色特征字典
            
        Returns:
            Dict[str, float]: 包含各项价值详情的字典
        """
        try:
            RULE = self.rule_setting
            
            # 读取市场折价系数配置
            rate_config_path = os.path.join(os.path.dirname(__file__), 'rate.jsonc')
            with open(rate_config_path, 'r', encoding='utf-8') as f:
                # 简单处理jsonc格式，去掉注释
                content = f.read()
                # 去掉行注释
                lines = content.split('\n')
                cleaned_lines = []
                for line in lines:
                    if '//' in line:
                        line = line[:line.index('//')]
                    cleaned_lines.append(line)
                cleaned_content = '\n'.join(cleaned_lines)
                DISCOUNT_RATES = json.loads(cleaned_content)
            
            value_breakdown = {}
            total_value = 0
            
            # 按照原始算法计算各项价值，然后应用折价系数
            
            # 1. 历史门派数加成
            school_history_count = features.get('school_history_count', 0)
            raw_value = school_history_count * RULE.get('ChangeSchoolCostPerTimes', 0)
            market_value = raw_value * DISCOUNT_RATES.get('school_history', 1.0)
            if raw_value > 0:
                value_breakdown['school_history'] = market_value
                total_value += market_value
                print(f"历史门派数: 原始{raw_value:.0f} -> 市场{market_value:.0f} (系数{DISCOUNT_RATES.get('school_history', 1.0)})")
            
            # 2. 乾元丹加成
            all_new_point = features.get('all_new_point', 0)
            qianyuandan_dict = RULE.get('QianyuandanCount2Value', {})
            if str(all_new_point) in qianyuandan_dict:
                raw_value = all_new_point * qianyuandan_dict[str(all_new_point)]
                market_value = raw_value * DISCOUNT_RATES.get('qianyuandan', 1.0)
                value_breakdown['qianyuandan'] = market_value
                total_value += market_value
                print(f"乾元丹: 原始{raw_value:.0f} -> 市场{market_value:.0f} (系数{DISCOUNT_RATES.get('qianyuandan', 1.0)})")
            
            # 3. 乾元丹突破加成
            if features.get('qianyuandan_breakthrough', False):
                raw_value = RULE.get('qianyuandanBreakthroughValue', 0)
                market_value = raw_value * DISCOUNT_RATES.get('qianyuandan', 1.0)  # 使用乾元丹的系数
                value_breakdown['qianyuandan_breakthrough'] = market_value
                total_value += market_value
                print(f"乾元丹突破: 原始{raw_value:.0f} -> 市场{market_value:.0f} (系数{DISCOUNT_RATES.get('qianyuandan', 1.0)})")
            
            # 4. 月饼粽子机缘
            jiyuan_amount = features.get('jiyuan_amount', 0)
            raw_value = jiyuan_amount * RULE.get('JiyuanValuePerCount', 0)
            market_value = raw_value * DISCOUNT_RATES.get('jiyuan', 1.0)
            if raw_value > 0:
                value_breakdown['jiyuan'] = market_value
                total_value += market_value
                print(f"月饼粽子机缘: 原始{raw_value:.0f} -> 市场{market_value:.0f} (系数{DISCOUNT_RATES.get('jiyuan', 1.0)})")
            
            # 5. 总经验
            sum_exp = features.get('sum_exp', 0)
            raw_value = sum_exp * RULE.get('ExpValuePerHundredMillion', 0)
            market_value = raw_value * DISCOUNT_RATES.get('exp', 1.0)
            if raw_value > 0:
                value_breakdown['exp'] = market_value
                total_value += market_value
                print(f"总经验: 原始{raw_value:.0f} -> 市场{market_value:.0f} (系数{DISCOUNT_RATES.get('exp', 1.0)})")
            
            # 6. 师门技能
            school_skills = features.get('school_skills', [])
            school_skill_dict = RULE.get('SchoolSkillGrade2Value', {})
            raw_value = 0
            for skill_grade in school_skills:
                if str(skill_grade) in school_skill_dict:
                    raw_value += school_skill_dict[str(skill_grade)]
            market_value = raw_value * DISCOUNT_RATES.get('school_skills', 1.0)
            if raw_value > 0:
                value_breakdown['school_skills'] = market_value
                total_value += market_value
                print(f"师门技能: 原始{raw_value:.0f} -> 市场{market_value:.0f} (系数{DISCOUNT_RATES.get('school_skills', 1.0)})")
            
            # 7. 行囊
            packet_page = features.get('packet_page', 0)
            packet_dict = RULE.get('PackagePage2Value', {})
            if str(packet_page) in packet_dict:
                raw_value = packet_dict[str(packet_page)]
                market_value = raw_value * DISCOUNT_RATES.get('packet', 1.0)
                value_breakdown['packet'] = market_value
                total_value += market_value
                print(f"行囊: 原始{raw_value:.0f} -> 市场{market_value:.0f} (系数{DISCOUNT_RATES.get('packet', 1.0)})")
            
            # 8. 储备金
            learn_cash = features.get('learn_cash', 0)
            raw_value = learn_cash/10000 * RULE.get('LearnCash2CashRate', 0)
            market_value = raw_value * DISCOUNT_RATES.get('learn_cash', 1.0)
            if raw_value > 0:
                value_breakdown['learn_cash'] = market_value
                total_value += market_value
                print(f"储备金: 原始{raw_value:.0f} -> 市场{market_value:.0f} (系数{DISCOUNT_RATES.get('learn_cash', 1.0)})")
            
            # 9. 仙玉
            xianyu_amount = features.get('xianyu_amount', 0)
            raw_value = xianyu_amount * RULE.get('RMB2MHB', 0)
            market_value = raw_value * DISCOUNT_RATES.get('xianyu', 1.0)
            if raw_value > 0:
                value_breakdown['xianyu'] = market_value
                total_value += market_value
                print(f"仙玉: 原始{raw_value:.0f} -> 市场{market_value:.0f} (系数{DISCOUNT_RATES.get('xianyu', 1.0)})")
            
            # 10. 修炼等级上限
            max_cultivate_dict = RULE.get('MaxCultivateGrade2Value', {})
            cultivate_cost_dict = RULE.get('CultivateCostPerCount', {})
            raw_value = 0
            for i in range(1, 5):
                max_expt_key = f'max_expt{i}'
                max_expt_value = features.get(max_expt_key, 0)
                cost_multiplier = cultivate_cost_dict.get(str(i), 0)
                if max_expt_value > 0 and str(max_expt_value) in max_cultivate_dict:
                    raw_value += max_cultivate_dict[str(max_expt_value)] * cost_multiplier
            market_value = raw_value * DISCOUNT_RATES.get('cultivation', 1.0)
            if raw_value > 0:
                value_breakdown['cultivation_limit'] = market_value
                total_value += market_value
                print(f"修炼等级上限: 原始{raw_value:.0f} -> 市场{market_value:.0f} (系数{DISCOUNT_RATES.get('cultivation', 1.0)})")
            
            # 11. 修炼等级
            cultivate_learn_count_dict = RULE['CultivateGrade2LearnCount']
            raw_value = 0
            for i in range(1, 5):
                expt_key = f'expt_ski{i}'
                expt_value = features.get(expt_key, 0)
                cost_multiplier = cultivate_cost_dict.get(str(i), 0)
                raw_value += cultivate_learn_count_dict[str(expt_value)] * cost_multiplier
            market_value = raw_value * DISCOUNT_RATES.get('cultivation', 1.0)
            if raw_value > 0:
                value_breakdown['cultivation_level'] = market_value
                total_value += market_value
                print(f"修炼等级: 原始{raw_value:.0f} -> 市场{market_value:.0f} (系数{DISCOUNT_RATES.get('cultivation', 1.0)})")
            
            # 12. 控制力
            control_dict = RULE.get('ControlGrade2XiulianguoCount', {})
            xiulianguo_value = RULE.get('XiulianguoValuePerCount', 0)
            raw_value = 0
            for i in range(1, 5):
                beast_ski_key = f'beast_ski{i}'
                beast_ski_value = features.get(beast_ski_key, 0)
                if str(beast_ski_value) in control_dict:
                    raw_value += xiulianguo_value * control_dict[str(beast_ski_value)]
            market_value = raw_value * DISCOUNT_RATES.get('beast_cultivation', 1.0)
            if raw_value > 0:
                value_breakdown['beast_cultivation'] = market_value
                total_value += market_value
                print(f"控制力: 原始{raw_value:.0f} -> 市场{market_value:.0f} (系数{DISCOUNT_RATES.get('beast_cultivation', 1.0)})")
            
            # 13. 生活技能
            life_skills = features.get('life_skills', [])
            life_skill_dict = RULE.get('LifeSkillGrade2Value', {})
            raw_value = 0
            for life_skill_grade in life_skills:
                if life_skill_grade >= 80 and str(life_skill_grade) in life_skill_dict:
                    raw_value += life_skill_dict[str(life_skill_grade)]
            market_value = raw_value * DISCOUNT_RATES.get('life_skills', 1.0)
            if raw_value > 0:
                value_breakdown['life_skills'] = market_value
                total_value += market_value
                print(f"生活技能: 原始{raw_value:.0f} -> 市场{market_value:.0f} (系数{DISCOUNT_RATES.get('life_skills', 1.0)})")
            
            # 14. 强壮、神速
            qiangzhuang_shensu = features.get('qiangzhuang&shensu', [])
            qiangzhuang_dict = RULE.get('QiangzhuangAndShensuGrade2Value', {})
            raw_value = 0
            for qiangzhuang_and_shensu_grade in qiangzhuang_shensu:
                if str(qiangzhuang_and_shensu_grade) in qiangzhuang_dict:
                    raw_value += qiangzhuang_dict[str(qiangzhuang_and_shensu_grade)]
            market_value = raw_value * DISCOUNT_RATES.get('qiangzhuang_shensu', 1.0)
            if raw_value > 0:
                value_breakdown['qiangzhuang_shensu'] = market_value
                total_value += market_value
                print(f"强壮、神速: 原始{raw_value:.0f} -> 市场{market_value:.0f} (系数{DISCOUNT_RATES.get('qiangzhuang_shensu', 1.0)})")
            
            # 15. 灵佑次数
            lingyou_count = features.get('lingyou_count', 0)
            raw_value = lingyou_count * RULE.get('LingyouValuePerCount', 0)
            market_value = raw_value * DISCOUNT_RATES.get('lingyou', 1.0)
            if raw_value > 0:
                value_breakdown['lingyou'] = market_value
                total_value += market_value
                print(f"灵佑次数: 原始{raw_value:.0f} -> 市场{market_value:.0f} (系数{DISCOUNT_RATES.get('lingyou', 1.0)})")
            
            # 16. 高成长坐骑数量
            hight_grow_rider_count = features.get('hight_grow_rider_count', 0)
            raw_value = hight_grow_rider_count * RULE.get('HightGrowRidePerValue', 0)
            market_value = raw_value * DISCOUNT_RATES.get('rider', 1.0)
            if raw_value > 0:
                value_breakdown['rider'] = market_value
                total_value += market_value
                print(f"高成长坐骑: 原始{raw_value:.0f} -> 市场{market_value:.0f} (系数{DISCOUNT_RATES.get('rider', 1.0)})")
            
            # 17. 召唤兽最大携带量
            allow_pet_count = features.get('allow_pet_count', 0)
            raw_value = RULE.get('AllowPetCount2Value', {}).get(str(allow_pet_count), 0)
            market_value = raw_value * DISCOUNT_RATES.get('pet_count', 1.0)
            if raw_value > 0:
                value_breakdown['pet_count'] = market_value
                total_value += market_value
                print(f"召唤兽携带量: 原始{raw_value:.0f} -> 市场{market_value:.0f} (系数{DISCOUNT_RATES.get('pet_count', 1.0)})")
            
            # 18. 有价值法宝数量
            premium_fabao_count = features.get('premium_fabao_count', 0)
            raw_value = premium_fabao_count * RULE.get('PremiumFabaoValuePerCount', 0)
            market_value = raw_value * DISCOUNT_RATES.get('fabao', 1.0)
            if raw_value > 0:
                value_breakdown['fabao'] = market_value
                total_value += market_value
                print(f"有价值法宝: 原始{raw_value:.0f} -> 市场{market_value:.0f} (系数{DISCOUNT_RATES.get('fabao', 1.0)})")
            
            # 19. 神器
            shenqi = features.get('shenqi', {})
            ShenqiAttr2Value = RULE.get('ShenqiAttr2Value', {})
            raw_value = 0
            for attr_key, multiplier_key in [
                ('same9Count', 'same9Count'),
                ('allTheSameCount', 'allTheSameCount'),
                ('1Count', '1'),
                ('1attrCount', '1attr'),
                ('2Count', '2'),
                ('2attrCount', '2attr'),
                ('3Count', '3'),
                ('3attrCount', '3attr')
            ]:
                if shenqi.get(attr_key, 0) > 0:
                    raw_value += shenqi[attr_key] * ShenqiAttr2Value.get(multiplier_key, 0)
            market_value = raw_value * DISCOUNT_RATES.get('shenqi', 1.0)
            if raw_value > 0:
                value_breakdown['shenqi'] = market_value
                total_value += market_value
                print(f"神器: 原始{raw_value:.0f} -> 市场{market_value:.0f} (系数{DISCOUNT_RATES.get('shenqi', 1.0)})")
            
            # 最终价值转换
            final_value = total_value / RULE['RMB2MHB'] * RULE['MARKET_FACTOR']
            
            print(f"\n=== 简化版价值分解 ===")
            for attr, value in value_breakdown.items():
                print(f"{attr}: {value:.0f}")
            print(f"总计: {total_value:.0f} -> 最终: {final_value:.0f}")
            
            return {
                'value_breakdown': value_breakdown,
                'total_raw_value': total_value,
                'final_value': final_value
            }
            
        except Exception as e:
            self.logger.error(f"简化版改进基础属性价值计算失败: {e}")
            import traceback
            self.logger.error(f"详细错误: {traceback.format_exc()}")
            return {
                'value_breakdown': {},
                'total_raw_value': 0.0,
                'final_value': 0.0
            }

    def export_to_excel_improved(self, output_path: str = 'results_improved.xlsx') -> Optional[str]:
        """
        导出改进算法的评估结果到Excel文件

        Args:
            output_path (str): 输出文件路径，默认为 'results_improved.xlsx'

        Returns:
            Optional[str]: 成功时返回输出文件路径，失败时返回None
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows

            if self.df is None or self.df.empty:
                print("没有数据可导出")
                return None

            print(f"开始导出 {len(self.df)} 条数据到Excel（改进算法）...")

            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "角色评估结果（改进算法）"

            # 准备数据
            export_data = []
            for equip_id, row in self.df.iterrows():
                # 获取角色基本信息
                price = row.get('price', 0)

                # 使用改进算法计算价值
                try:
                    improved_result = self.calc_base_attributes_value_improved(row.to_dict())
                    improved_value = improved_result['final_value']
                    value_breakdown = improved_result['value_breakdown']
                except:
                    improved_value = 0
                    value_breakdown = {}

                # 生成CBG链接
                cbg_link = self.generate_cbg_link(equip_id)

                # 准备一行数据
                row_data = {
                    'equip_id': equip_id,
                    'cbg_link': cbg_link,
                    'price': price,
                    'improved_value': improved_value,
                    'price_diff_ratio': (improved_value - price) / price * 100 if price > 0 else 0,
                    **value_breakdown,  # 展开价值分解
                    **row.to_dict()  # 包含所有features
                }
                export_data.append(row_data)

            if not export_data:
                print("没有有效数据可导出")
                return None

            # 定义列顺序：优先列 + 价值分解列 + features列
            priority_columns = ['equip_id', 'price', 'improved_value', 'price_diff_ratio']
            
            # 获取所有价值分解列
            value_breakdown_columns = []
            for data in export_data:
                for key in data.keys():
                    if key not in priority_columns and key not in ['cbg_link'] and key not in export_data[0].keys():
                        continue
                    # 检查是否是价值分解列（通常是字符串key且不在features中）
                    if (key in ['school_history', 'jiyuan', 'experience', 'learn_cash', 'xianyu', 
                               'lingyou', 'rider', 'fabao', 'qianyuandan', 'packet', 'pet_count',
                               'qianyuandan_breakthrough', 'school_skills', 'cultivation_limit',
                               'cultivation_level', 'beast_cultivation', 'life_skills', 'qiangzhuang_shensu', 'shenqi']):
                        if key not in value_breakdown_columns:
                            value_breakdown_columns.append(key)

            # 获取所有features列（排除优先列和价值分解列）
            all_columns = list(export_data[0].keys())
            feature_columns = [
                col for col in all_columns 
                if col not in priority_columns and col != 'cbg_link' and col not in value_breakdown_columns
            ]

            # 最终列顺序
            final_columns = priority_columns + value_breakdown_columns + feature_columns

            # 设置表头
            headers = []
            for col in final_columns:
                if col == 'equip_id':
                    headers.append('装备ID')
                elif col == 'price':
                    headers.append('价格')
                elif col == 'improved_value':
                    headers.append('改进估值')
                elif col == 'price_diff_ratio':
                    headers.append('差价比率(%)')
                elif col in value_breakdown_columns:
                    headers.append(f'{col}_价值')
                else:
                    headers.append(col)

            # 写入表头
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(
                    horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

            # 写入数据
            for row_idx, data in enumerate(export_data, 2):
                for col_idx, col_name in enumerate(final_columns, 1):
                    value = data.get(col_name, '')

                    # 特殊处理角色名列（添加超链接）
                    if col_name == 'equip_id':
                        cbg_link = data.get('cbg_link')
                        cell = ws.cell(
                            row=row_idx, column=col_idx, value=value)
                        if cbg_link:
                            cell.hyperlink = cbg_link
                            cell.font = Font(
                                color="0000FF", underline="single")
                        else:
                            cell.font = Font()
                    else:
                        # 数值格式化
                        if isinstance(value, (int, float)):
                            if col_name in ['price', 'improved_value', 'price_diff_ratio'] or col_name in value_breakdown_columns:
                                # 价格显示
                                display_value = f"{value:.1f}" if value > 0 else "0"
                            elif '_ratio' in col_name:
                                # 比率显示
                                display_value = f"{value:.1f}%" if value != 0 else "0%"
                            else:
                                display_value = f"{value:.2f}" if isinstance(
                                    value, float) else str(value)
                        else:
                            display_value = str(
                                value) if value is not None else ""

                        cell = ws.cell(
                            row=row_idx, column=col_idx, value=display_value)

                    # 设置单元格样式
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center")
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )

                    # 特殊颜色
                    if col_name == 'price':
                        cell.fill = PatternFill(
                            start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                    elif col_name == 'improved_value':
                        cell.fill = PatternFill(
                            start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                    elif col_name == 'price_diff_ratio':
                        # 根据差价比率设置颜色
                        value = float(cell.value) if cell.value != '' else 0
                        if value > 20:  # 高溢价
                            cell.fill = PatternFill(
                                start_color="FF0000", end_color="FF0000", fill_type="solid")
                            cell.font = Font(color="FFFFFF")
                        elif value < -20:  # 高折价
                            cell.fill = PatternFill(
                                start_color="00FF00", end_color="00FF00", fill_type="solid")
                            cell.font = Font(color="000000")
                    elif col_name in value_breakdown_columns:
                        # 价值分解列使用浅蓝色
                        cell.fill = PatternFill(
                            start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")

            # 调整列宽
            for col_idx, col_name in enumerate(final_columns, 1):
                if col_name in ['equip_id']:
                    ws.column_dimensions[openpyxl.utils.get_column_letter(
                        col_idx)].width = 25
                elif col_name in ['price', 'improved_value', 'price_diff_ratio'] or col_name in value_breakdown_columns:
                    ws.column_dimensions[openpyxl.utils.get_column_letter(
                        col_idx)].width = 15
                elif '_ratio' in col_name:
                    ws.column_dimensions[openpyxl.utils.get_column_letter(
                        col_idx)].width = 18
                else:
                    ws.column_dimensions[openpyxl.utils.get_column_letter(
                        col_idx)].width = 12

            # 冻结首行
            ws.freeze_panes = "A2"

            # 添加算法对比工作表
            comparison_ws = wb.create_sheet("算法对比统计")
            comparison_data = [
                ["统计项目", "改进算法", "说明"],
                ["总角色数", len(export_data), ""],
                ["平均估值", f"{sum(d['improved_value'] for d in export_data)/len(export_data):.2f}", "改进算法平均估值"],
                ["价值分解项数", len(value_breakdown_columns), "包含的价值分解维度"],
                ["折价系数应用", "是", "已应用市场折价系数"],
                ["成本类型区分", "线性/阶梯/特殊", "三种成本类型"],
            ]

            for row_idx, row_data in enumerate(comparison_data, 1):
                for col_idx, value in enumerate(row_data, 1):
                    cell = comparison_ws.cell(
                        row=row_idx, column=col_idx, value=value)
                    if row_idx == 1:  # 表头
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(
                            start_color="366092", end_color="366092", fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")

            # 保存文件
            wb.save(output_path)
            print(f"Excel文件已保存到: {output_path}")
            print(f"包含 {len(export_data)} 条角色数据")
            print(f"优先列: {priority_columns}")
            print(f"价值分解列: {value_breakdown_columns}")
            print(f"特征列数: {len(feature_columns)}")

            return output_path

        except ImportError:
            print("需要安装openpyxl库: pip install openpyxl")
            return None
        except Exception as e:
            print(f"导出Excel失败: {e}")
            import traceback
            traceback.print_exc()
            return None

    def export_comparison_to_excel(self, output_path: str = 'results_comparison.xlsx') -> Optional[str]:
        """
        导出原始算法和改进算法的对比结果到Excel文件

        Args:
            output_path (str): 输出文件路径，默认为 'results_comparison.xlsx'

        Returns:
            Optional[str]: 成功时返回输出文件路径，失败时返回None
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows

            if self.df is None or self.df.empty:
                print("没有数据可导出")
                return None

            print(f"开始导出 {len(self.df)} 条数据到Excel（算法对比）...")

            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "算法对比结果"

            # 准备数据
            export_data = []
            for equip_id, row in self.df.iterrows():
                # 获取角色基本信息
                price = row.get('price', 0)

                # 计算原始算法价值
                try:
                    original_value = self.calc_base_attributes_value(row.to_dict())
                except:
                    original_value = 0

                # 计算改进算法价值
                try:
                    improved_result = self.calc_base_attributes_value_improved(row.to_dict())
                    improved_value = improved_result['final_value']
                    value_breakdown = improved_result['value_breakdown']
                except:
                    improved_value = 0
                    value_breakdown = {}

                # 生成CBG链接
                cbg_link = self.generate_cbg_link(equip_id)

                # 计算差异
                original_diff_ratio = (original_value - price) / price * 100 if price > 0 else 0
                improved_diff_ratio = (improved_value - price) / price * 100 if price > 0 else 0
                algorithm_diff = improved_value - original_value
                algorithm_diff_ratio = (algorithm_diff / original_value * 100) if original_value > 0 else 0

                # 准备一行数据
                row_data = {
                    'equip_id': equip_id,
                    'cbg_link': cbg_link,
                    'price': price,
                    'original_value': original_value,
                    'improved_value': improved_value,
                    'original_diff_ratio': original_diff_ratio,
                    'improved_diff_ratio': improved_diff_ratio,
                    'algorithm_diff': algorithm_diff,
                    'algorithm_diff_ratio': algorithm_diff_ratio,
                    **value_breakdown,  # 展开价值分解
                    **row.to_dict()  # 包含所有features
                }
                export_data.append(row_data)

            if not export_data:
                print("没有有效数据可导出")
                return None

            # 定义列顺序：优先列 + 对比列 + 价值分解列 + features列
            priority_columns = ['equip_id', 'price', 'original_value', 'improved_value', 
                              'original_diff_ratio', 'improved_diff_ratio', 'algorithm_diff', 'algorithm_diff_ratio']
            
            # 获取所有价值分解列
            value_breakdown_columns = []
            for data in export_data:
                for key in data.keys():
                    # 检查是否是价值分解列
                    if (key in ['school_history', 'jiyuan', 'exp', 'learn_cash', 'xianyu', 
                               'lingyou', 'rider', 'fabao', 'qianyuandan', 'packet', 'pet_count',
                               'qianyuandan_breakthrough', 'school_skills', 'cultivation_limit',
                               'cultivation_level', 'beast_cultivation', 'life_skills', 'qiangzhuang_shensu', 'shenqi']):
                        if key not in value_breakdown_columns:
                            value_breakdown_columns.append(key)

            # 获取所有features列（排除优先列和价值分解列）
            all_columns = list(export_data[0].keys())
            feature_columns = [
                col for col in all_columns 
                if col not in priority_columns and col != 'cbg_link' and col not in value_breakdown_columns
            ]

            # 最终列顺序
            final_columns = priority_columns + value_breakdown_columns + feature_columns

            # 设置表头
            headers = []
            for col in final_columns:
                if col == 'equip_id':
                    headers.append('装备ID')
                elif col == 'price':
                    headers.append('价格')
                elif col == 'original_value':
                    headers.append('原始估值')
                elif col == 'improved_value':
                    headers.append('改进估值')
                elif col == 'original_diff_ratio':
                    headers.append('原始差价比率(%)')
                elif col == 'improved_diff_ratio':
                    headers.append('改进差价比率(%)')
                elif col == 'algorithm_diff':
                    headers.append('算法差异')
                elif col == 'algorithm_diff_ratio':
                    headers.append('算法差异比率(%)')
                elif col in value_breakdown_columns:
                    headers.append(f'{col}_价值')
                else:
                    headers.append(col)

            # 写入表头
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(
                    horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

            # 写入数据
            for row_idx, data in enumerate(export_data, 2):
                for col_idx, col_name in enumerate(final_columns, 1):
                    value = data.get(col_name, '')

                    # 特殊处理角色名列（添加超链接）
                    if col_name == 'equip_id':
                        cbg_link = data.get('cbg_link')
                        cell = ws.cell(
                            row=row_idx, column=col_idx, value=value)
                        if cbg_link:
                            cell.hyperlink = cbg_link
                            cell.font = Font(
                                color="0000FF", underline="single")
                        else:
                            cell.font = Font()
                    else:
                        # 数值格式化
                        if isinstance(value, (int, float)):
                            if col_name in ['price', 'original_value', 'improved_value', 'algorithm_diff'] or col_name in value_breakdown_columns:
                                # 价格显示
                                display_value = f"{value:.1f}" if value > 0 else "0"
                            elif '_ratio' in col_name:
                                # 比率显示
                                display_value = f"{value:.1f}%" if value != 0 else "0%"
                            else:
                                display_value = f"{value:.2f}" if isinstance(
                                    value, float) else str(value)
                        else:
                            display_value = str(
                                value) if value is not None else ""

                        cell = ws.cell(
                            row=row_idx, column=col_idx, value=display_value)

                    # 设置单元格样式
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center")
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )

                    # 特殊颜色
                    if col_name == 'price':
                        cell.fill = PatternFill(
                            start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                    elif col_name == 'original_value':
                        cell.fill = PatternFill(
                            start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                    elif col_name == 'improved_value':
                        cell.fill = PatternFill(
                            start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                    elif col_name in ['original_diff_ratio', 'improved_diff_ratio']:
                        # 根据差价比率设置颜色
                        value = float(data.get(col_name, 0))
                        if value > 20:  # 高溢价
                            cell.fill = PatternFill(
                                start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                        elif value < -20:  # 高折价
                            cell.fill = PatternFill(
                                start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                    elif col_name == 'algorithm_diff':
                        # 算法差异着色
                        value = float(data.get(col_name, 0))
                        if value > 0:
                            cell.fill = PatternFill(
                                start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                        elif value < 0:
                            cell.fill = PatternFill(
                                start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")
                    elif col_name == 'algorithm_diff_ratio':
                        # 根据算法差异比率设置颜色
                        value = float(data.get(col_name, 0))
                        if abs(value) > 10:  # 差异较大
                            cell.fill = PatternFill(
                                start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
                    elif col_name in value_breakdown_columns:
                        # 价值分解列使用浅蓝色
                        cell.fill = PatternFill(
                            start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")

            # 调整列宽
            for col_idx, col_name in enumerate(final_columns, 1):
                if col_name in ['equip_id']:
                    ws.column_dimensions[openpyxl.utils.get_column_letter(
                        col_idx)].width = 25
                elif col_name in ['price', 'original_value', 'improved_value', 'algorithm_diff'] or col_name in value_breakdown_columns:
                    ws.column_dimensions[openpyxl.utils.get_column_letter(
                        col_idx)].width = 15
                elif '_ratio' in col_name:
                    ws.column_dimensions[openpyxl.utils.get_column_letter(
                        col_idx)].width = 18
                else:
                    ws.column_dimensions[openpyxl.utils.get_column_letter(
                        col_idx)].width = 12

            # 冻结首行
            ws.freeze_panes = "A2"

            # 添加算法对比统计工作表
            comparison_ws = wb.create_sheet("算法对比统计")
            
            # 计算统计数据
            original_values = [d['original_value'] for d in export_data]
            improved_values = [d['improved_value'] for d in export_data]
            prices = [d['price'] for d in export_data]
            
            original_avg = sum(original_values) / len(original_values)
            improved_avg = sum(improved_values) / len(improved_values)
            price_avg = sum(prices) / len(prices)
            
            # 计算原始算法准确率（差价比率在±20%内的比例）
            original_accurate = sum(1 for d in export_data if abs(d['original_diff_ratio']) <= 20) / len(export_data) * 100
            improved_accurate = sum(1 for d in export_data if abs(d['improved_diff_ratio']) <= 20) / len(export_data) * 100
            
            comparison_data = [
                ["统计项目", "原始算法", "改进算法", "差异", "说明"],
                ["总角色数", len(export_data), len(export_data), 0, ""],
                ["平均估值", f"{original_avg:.1f}", f"{improved_avg:.1f}", f"{improved_avg-original_avg:.1f}", ""],
                ["平均实际价格", f"{price_avg:.1f}", f"{price_avg:.1f}", 0, ""],
                ["估值准确率(±20%)", f"{original_accurate:.1f}%", f"{improved_accurate:.1f}%", f"{improved_accurate-original_accurate:.1f}%", "差价比率在±20%内的比例"],
                ["价值分解项数", "无", len(value_breakdown_columns), len(value_breakdown_columns), "改进算法包含的价值分解维度"],
                ["折价系数应用", "否", "是", "是", "是否应用市场折价系数"],
            ]

            for row_idx, row_data in enumerate(comparison_data, 1):
                for col_idx, value in enumerate(row_data, 1):
                    cell = comparison_ws.cell(
                        row=row_idx, column=col_idx, value=value)
                    if row_idx == 1:  # 表头
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(
                            start_color="366092", end_color="366092", fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )

            # 调整统计表列宽
            for col_idx in range(1, 6):
                comparison_ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

            # 保存文件
            wb.save(output_path)
            print(f"Excel对比文件已保存到: {output_path}")
            print(f"包含 {len(export_data)} 条角色数据")
            print(f"对比列: 原始估值 vs 改进估值")
            print(f"价值分解列: {value_breakdown_columns}")
            print(f"原始算法准确率: {original_accurate:.1f}%")
            print(f"改进算法准确率: {improved_accurate:.1f}%")

            return output_path

        except ImportError:
            print("需要安装openpyxl库: pip install openpyxl")
            return None
        except Exception as e:
            print(f"导出Excel对比文件失败: {e}")
            import traceback
            traceback.print_exc()
            return None


if __name__ == "__main__":
    try:
        # 自动查找/data/文件夹下的所有.db文件
        db_files = glob.glob("data/*.db")
        if not db_files:
            print("未找到数据库文件，请确保/data/文件夹下有.db文件")
            exit(1)
        # 选择第一个.db文件
        db_path = db_files[0]
        print(f"使用数据库文件: {db_path}")

        evaluator = CharacterEvaluator()
        print("开始加载数据...")
        # 加载数据
        evaluator.fit(db_path)
        print("数据加载完成")

        print("\n=== 导出原始算法和改进算法的对比结果 ===")
        # 导出原始算法和改进算法的对比结果到Excel
        comparison_output_path = evaluator.export_comparison_to_excel(
            './output/results_comparison.xlsx')
        if comparison_output_path:
            print(f"算法对比结果已成功导出到: {comparison_output_path}")
        else:
            print("导出算法对比结果Excel失败")

    except Exception as e:
        print(f"程序运行出错: {str(e)}")
        import traceback
        print(traceback.format_exc())
